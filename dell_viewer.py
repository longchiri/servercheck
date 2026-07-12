#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Dell iDRAC Viewer (PySide6, cross-platform)
- macOS / Windows 모두 동일 코드로 동작
- HW / FW / BIOS / 전체 조회
- 엑셀(.xlsx) 누적 저장 (서버별 시트 추가)
- 보기 좋게 토글, 폰트 조절, 샘플 보기, 친절한 에러 메시지
"""
from __future__ import annotations

import os
import sys
import json
import html
import base64
import warnings
import traceback
from datetime import datetime
from typing import Callable, Optional, List, Tuple

import requests
from requests.auth import HTTPBasicAuth

from PySide6.QtCore import Qt, QThread, Signal, QSettings, QUrl
from PySide6.QtGui import QFont, QFontDatabase, QAction, QKeySequence, QGuiApplication, QIcon, QPixmap, QDesktopServices
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QLineEdit, QPushButton,
    QComboBox, QSpinBox, QTextEdit, QTextBrowser, QVBoxLayout, QHBoxLayout,
    QGridLayout, QGroupBox, QMessageBox, QFileDialog, QCheckBox, QStatusBar,
    QSizePolicy, QFrame, QDialog, QProgressBar, QTabWidget, QScrollArea,
)


def resource_path(name: str) -> str:
    """PyInstaller 번들/일반 실행 모두에서 동일하게 리소스 찾기"""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    p = os.path.join(base, name)
    if os.path.exists(p):
        return p
    # 개발 시 .py 와 같은 폴더
    p2 = os.path.join(os.path.dirname(os.path.abspath(__file__)), name)
    return p2 if os.path.exists(p2) else ""


# 앱 전역 스타일시트 — 깔끔한 카드/버튼 톤
APP_STYLE = """
QMainWindow, QWidget { background: #f6f8fa; color: #1f2328; }
QGroupBox {
    background: #ffffff;
    border: 1px solid #d0d7de;
    border-radius: 8px;
    margin-top: 14px;
    padding-top: 18px;
    font-weight: 600;
}
QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left;
    left: 14px;
    padding: 0 6px;
    color: #57606a;
    background: transparent;
}
QLabel { background: transparent; color: #57606a; font-size: 12px; }
QLineEdit, QSpinBox, QComboBox {
    background: #ffffff;
    border: 1px solid #d0d7de;
    border-radius: 6px;
    padding: 6px 10px;
    selection-background-color: #0969da;
    selection-color: white;
    font-size: 13px;
    color: #1f2328;
}
QLineEdit:focus, QSpinBox:focus, QComboBox:focus { border: 1px solid #0969da; }
QLineEdit:disabled { background: #f6f8fa; color: #8c959f; }
QPushButton {
    background: #ffffff;
    border: 1px solid #d0d7de;
    border-radius: 6px;
    padding: 6px 14px;
    color: #1f2328;
    font-size: 12px;
}
QPushButton:hover { background: #f3f4f6; border-color: #afb8c1; }
QPushButton:pressed { background: #ebecef; }
QPushButton:disabled { background: #f6f8fa; color: #8c959f; border-color: #e1e4e8; }
QCheckBox { color: #1f2328; font-size: 13px; spacing: 6px; }
QCheckBox::indicator { width: 16px; height: 16px; }
QStatusBar { background: #ffffff; border-top: 1px solid #d0d7de; color: #57606a; }
QTextBrowser {
    background: #ffffff;
    border: 1px solid #d0d7de;
    border-radius: 8px;
    padding: 12px;
    selection-background-color: #ddf4ff;
    selection-color: #1f2328;
}
"""

# ----------- SSL 경고 끄기 ------------
warnings.filterwarnings("ignore")
try:
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
except Exception:
    pass


APP_NAME = "iDRAC Toolkit"
APP_VERSION = "3.1.0"
ISSUES_URL = "https://github.com/longchiri/servercheck/issues/new"

# =========================================================
#  iDRAC Redfish Inspector
# =========================================================
class InspectorError(Exception):
    pass


class Inspector:
    """elgfw.py 로직을 옮긴 핵심 조회 클래스. UI 와 분리."""

    def __init__(self, ip: str, user: str, password: str, timeout: int = 15):
        self.ip = ip
        self.user = user
        self.password = password
        self.timeout = timeout
        self.session = requests.Session()
        self.session.verify = False
        # ✅ UTF-8 호환 Basic Auth (한글 비밀번호, 특수문자 등 latin-1 외 문자 지원)
        # 기본 HTTPBasicAuth 는 latin-1 만 처리해서 UnicodeEncodeError 발생 →
        # 직접 base64 인코딩한 Authorization 헤더로 우회
        creds = f"{user}:{password}".encode("utf-8")
        self.session.headers["Authorization"] = "Basic " + base64.b64encode(creds).decode("ascii")

    def _url(self, path: str) -> str:
        return f"https://{self.ip}{path}"

    def fetch_service_tag(self) -> str:
        """Service Tag 만 빠르게 조회"""
        try:
            d = self._get("/redfish/v1/Systems/System.Embedded.1")
            return d.get("SKU", "N/A") or "N/A"
        except Exception:
            return "N/A"

    # ---------- 로그 (LCLog / SEL) ----------
    LOG_PATHS = {
        "lclog": "/redfish/v1/Managers/iDRAC.Embedded.1/LogServices/Lclog/Entries",
        "sel":   "/redfish/v1/Managers/iDRAC.Embedded.1/LogServices/Sel/Entries",
    }
    LOG_NAMES = {
        "lclog": "Lifecycle Controller Log",
        "sel":   "System Event Log (SEL)",
    }

    def fetch_log_count(self, log_type: str) -> int:
        """전체 entry 개수 조회 (진행률 계산용)"""
        path = self.LOG_PATHS.get(log_type)
        if not path:
            return 0
        # $top=1 로 첫 1개만 받아서 총 개수 확인
        try:
            d = self._get(f"{path}?$top=1")
            return int(d.get("Members@odata.count")
                       or d.get("@odata.count") or 0)
        except Exception:
            return 0

    def fetch_log_page(self, log_type: str, skip: int, top: int = 50) -> list:
        """페이지 단위로 entry 조회"""
        path = self.LOG_PATHS.get(log_type)
        if not path:
            return []
        try:
            d = self._get(f"{path}?$skip={skip}&$top={top}")
            return d.get("Members", []) or []
        except Exception:
            return []

    # ---------- 펌웨어 업데이트 ----------
    def get_update_service(self) -> dict:
        """UpdateService 정보 조회 (지원 방식 등)"""
        return self._get("/redfish/v1/UpdateService")

    def check_update_ready(self) -> Tuple[bool, str]:
        """iDRAC 상태 검증 — 진행 중 Job 있는지, 정상 상태인지"""
        try:
            # 진행 중인 Job 확인
            jobs = self._get("/redfish/v1/Managers/iDRAC.Embedded.1/Oem/Dell/Jobs")
            running = []
            for m in jobs.get("Members", []):
                try:
                    j = self._get(m["@odata.id"])
                    state = j.get("JobState", "")
                    if state in ("Running", "Downloading", "Downloaded", "Scheduling"):
                        running.append(f"{j.get('Id', '?')} ({j.get('Name', '')}: {state})")
                except InspectorError:
                    pass
            if running:
                return False, "진행 중인 Job 이 있습니다:\n  " + "\n  ".join(running)
            return True, "OK"
        except InspectorError as e:
            return True, f"(상태 확인 실패, 계속 진행 가능: {e})"

    def _extract_error_detail(self, r) -> str:
        """iDRAC Redfish 에러 응답에서 상세 사유 추출"""
        try:
            body = r.json()
            err = body.get("error", body)
            if isinstance(err, dict):
                ext = err.get("@Message.ExtendedInfo", [])
                if isinstance(ext, list) and ext:
                    msgs = []
                    for m in ext[:3]:
                        mid = m.get("MessageId", "")
                        msg = m.get("Message", "")
                        reso = m.get("Resolution", "")
                        line = f"  · [{mid}] {msg}"
                        if reso and reso != "No response action is required.":
                            line += f"\n    → {reso}"
                        msgs.append(line)
                    return "\n" + "\n".join(msgs)
                elif err.get("message"):
                    return f"\n  · {err.get('message')}"
        except Exception:
            pass
        return f"\n  응답: {r.text[:400]}"

    def upload_firmware_multipart(self, file_path: str, apply_time: str = "Immediate",
                                   progress_callback=None) -> str:
        """Multipart HTTP Push 로 펌웨어 파일 업로드.
           - 우선 Dell 공식 예제 스타일의 MultipartUpload 시도
           - 실패 시 HttpPushUri 방식 폴백
           반환: Job/Task URI
        """
        import mimetypes
        from requests_toolbelt.multipart.encoder import MultipartEncoder, MultipartEncoderMonitor

        # UpdateService 에서 정확한 URI 조회
        multipart_uri = None
        http_push_uri = None
        try:
            us = self._get("/redfish/v1/UpdateService")
            multipart_uri = us.get("MultipartHttpPushUri")
            http_push_uri = us.get("HttpPushUri")
        except Exception:
            pass
        # 기본값 (표준 Redfish 경로)
        if not multipart_uri:
            multipart_uri = "/redfish/v1/UpdateService/MultipartUpload"

        filename = os.path.basename(file_path)

        # ============================================================
        # 방식 1: MultipartHttpPushUri (Redfish 표준, 권장)
        # ============================================================
        upload_url = self._url(multipart_uri) if multipart_uri.startswith("/") else multipart_uri

        # ⚠️ Dell 공식 예제와 정확히 일치하도록:
        # - UpdateParameters 파트의 파일명 = None (Dell iDRAC 이 파일명 있으면 거부)
        # - UpdateFile 파트의 Content-Type = "multipart/form-data" (Dell 명세)
        params = {
            "Targets": [],
            "@Redfish.OperationApplyTime": apply_time,
            "Oem": {},
        }

        file_handle = open(file_path, "rb")
        multipart_error = None
        r = None
        try:
            encoder = MultipartEncoder(
                fields={
                    "UpdateParameters": (None, json.dumps(params), "application/json"),
                    "UpdateFile": (filename, file_handle, "multipart/form-data"),
                }
            )

            def cb(monitor):
                if progress_callback:
                    progress_callback(monitor.bytes_read, monitor.len)

            monitor = MultipartEncoderMonitor(encoder, cb)

            r = requests.post(
                upload_url,
                data=monitor,
                headers={
                    "Content-Type": monitor.content_type,
                    "Authorization": self.session.headers["Authorization"],
                    "Accept": "application/json",
                },
                verify=False,
                timeout=1800,
            )

            if r.status_code in (200, 201, 202):
                # 성공 — Job URI 추출
                task_uri = r.headers.get("Location", "")
                if not task_uri:
                    try:
                        body = r.json()
                        task_uri = (body.get("@odata.id")
                                    or body.get("TaskMonitor")
                                    or body.get("Id", ""))
                    except Exception:
                        pass
                if not task_uri:
                    raise InspectorError("업로드 성공했지만 Job/Task URI를 못 찾음")
                return task_uri
            else:
                multipart_error = (
                    f"HTTP {r.status_code}\n"
                    f"URL: {upload_url}\n"
                    f"응답:{self._extract_error_detail(r)}"
                )
        except InspectorError:
            raise
        except Exception as e:
            multipart_error = f"예외: {e}"
        finally:
            try:
                file_handle.close()
            except Exception:
                pass

        # ============================================================
        # 방식 2: HttpPushUri 폴백 (iDRAC 8 또는 구형 iDRAC 9)
        # ============================================================
        if http_push_uri:
            try:
                push_url = self._url(http_push_uri) if http_push_uri.startswith("/") else http_push_uri
                # ETag 획득 (필수)
                etag = ""
                try:
                    head = requests.get(push_url, headers={
                        "Authorization": self.session.headers["Authorization"]
                    }, verify=False, timeout=30)
                    etag = head.headers.get("ETag", "")
                except Exception:
                    pass

                # 파일 전체 로드 (스트리밍보다 안전)
                with open(file_path, "rb") as f:
                    file_bytes = f.read()

                headers = {
                    "Content-Type": "multipart/form-data",
                    "Authorization": self.session.headers["Authorization"],
                    "Accept": "application/json",
                }
                if etag:
                    headers["If-Match"] = etag

                # 진행률은 파일 로드 완료로 처리 (스트리밍 폴백)
                if progress_callback:
                    progress_callback(len(file_bytes), len(file_bytes))

                r2 = requests.post(
                    push_url,
                    data=file_bytes,
                    headers=headers,
                    verify=False,
                    timeout=1800,
                )
                if r2.status_code in (200, 201, 202):
                    task_uri = r2.headers.get("Location", "")
                    if not task_uri:
                        try:
                            body = r2.json()
                            task_uri = (body.get("@odata.id")
                                        or body.get("TaskMonitor")
                                        or body.get("Id", ""))
                        except Exception:
                            pass
                    if task_uri:
                        return task_uri
                    else:
                        raise InspectorError("HttpPush 성공했지만 Job URI를 못 찾음")
                # 폴백도 실패
                fallback_error = (
                    f"[HttpPushUri 폴백도 실패]\n"
                    f"HTTP {r2.status_code}\n"
                    f"응답:{self._extract_error_detail(r2)}"
                )
                raise InspectorError(
                    f"펌웨어 업로드 실패\n\n"
                    f"[방식 1 MultipartUpload]\n{multipart_error}\n\n"
                    f"{fallback_error}"
                )
            except InspectorError:
                raise
            except Exception as e:
                raise InspectorError(
                    f"펌웨어 업로드 실패\n\n"
                    f"[방식 1 MultipartUpload]\n{multipart_error}\n\n"
                    f"[방식 2 HttpPushUri 예외]\n{e}"
                )

        # 방식 2 자체가 없음 → 방식 1 에러 그대로
        raise InspectorError(f"펌웨어 업로드 실패\n{multipart_error}")

    def poll_job(self, job_uri: str) -> dict:
        """Job/Task 상태 조회. 반환: {percent, state, messages}"""
        try:
            d = self._get(job_uri)
        except InspectorError as e:
            return {"percent": 0, "state": "Unknown", "message": str(e), "raw": {}}
        # Dell OEM Job
        pct = d.get("PercentComplete")
        if pct is None:
            pct = d.get("TaskProgress")
        state = d.get("JobState") or d.get("TaskState") or "Unknown"
        messages = d.get("Messages", [])
        msg = ""
        if messages:
            msg = messages[-1].get("Message", "") if isinstance(messages[-1], dict) else str(messages[-1])
        elif d.get("Message"):
            msg = d.get("Message", "")
        return {
            "percent": int(pct) if pct is not None else 0,
            "state": state,
            "message": msg,
            "raw": d,
        }

    def _get(self, path_or_url: str) -> dict:
        url = path_or_url if path_or_url.startswith("http") else self._url(path_or_url)
        r = self.session.get(url, timeout=self.timeout)
        if r.status_code == 401:
            raise InspectorError("AUTH")  # 인증 실패
        if r.status_code == 404:
            return {}  # 없음
        if r.status_code != 200:
            raise InspectorError(f"HTTP {r.status_code}: {url}")
        try:
            return r.json()
        except json.JSONDecodeError:
            raise InspectorError("JSON 파싱 실패")

    # ---------- HW ----------
    def fetch_hardware(self) -> dict:
        data = self._get("/redfish/v1/Systems/System.Embedded.1")

        result = {
            "system": {
                "Model": data.get("Model", "N/A"),
                "Manufacturer": data.get("Manufacturer", "N/A"),
                "ServiceTag": data.get("SKU", "N/A"),
                "PowerState": data.get("PowerState", "N/A"),
                "Health": data.get("Status", {}).get("Health", "N/A"),
            },
        }

        # CPU summary
        ps = data.get("ProcessorSummary", {})
        result["cpu_summary"] = {
            "Model": ps.get("Model", "N/A"),
            "Count": ps.get("Count", "N/A"),
            "Health": ps.get("Status", {}).get("Health", "N/A"),
        }
        # CPU 상세
        cpus = []
        try:
            proc_list = self._get("/redfish/v1/Systems/System.Embedded.1/Processors")
            for m in proc_list.get("Members", []):
                d = self._get(m["@odata.id"])
                cpus.append({
                    "Socket": d.get("Socket", "N/A"),
                    "Speed": d.get("OperatingSpeedMHz", "N/A"),
                    "Cores": d.get("TotalCores", "N/A"),
                })
        except InspectorError:
            pass
        result["cpus"] = cpus

        # 메모리
        ms = data.get("MemorySummary", {})
        result["memory_summary"] = {
            "TotalGiB": ms.get("TotalSystemMemoryGiB", "N/A"),
            "Health": ms.get("Status", {}).get("Health", "N/A"),
        }
        mems = []
        try:
            mem_list = self._get("/redfish/v1/Systems/System.Embedded.1/Memory")
            for m in mem_list.get("Members", []):
                d = self._get(m["@odata.id"])
                cap_mib = d.get("CapacityMiB")
                cap_gb = round(cap_mib / 1024, 2) if isinstance(cap_mib, (int, float)) else "N/A"
                mems.append({
                    "Location": d.get("DeviceLocator", "N/A"),
                    "CapacityGB": cap_gb,
                    "SpeedMHz": d.get("OperatingSpeedMhz", "N/A"),
                })
        except InspectorError:
            pass
        result["memory_modules"] = mems

        # 팬
        fans = []
        try:
            th = self._get("/redfish/v1/Chassis/System.Embedded.1/Thermal")
            for f in th.get("Fans", []):
                fans.append({
                    "Name": f.get("FanName") or f.get("Name", "N/A"),
                    "Health": f.get("Status", {}).get("Health", "N/A"),
                })
        except InspectorError:
            pass
        result["fans"] = fans

        # 스토리지
        controllers = []
        raids = []
        disks = []
        try:
            st = self._get("/redfish/v1/Systems/System.Embedded.1/Storage")
            for m in st.get("Members", []):
                d = self._get(m["@odata.id"])
                cname = d.get("Name", "N/A")
                chealth = d.get("Status", {}).get("Health", "N/A")
                if "PERC" in cname or "BOSS" in cname:
                    controllers.append({"Name": cname, "Health": chealth})

                vol_link = d.get("Volumes", {}).get("@odata.id")
                if vol_link:
                    try:
                        vols = self._get(vol_link)
                        for v in vols.get("Members", []):
                            vd = self._get(v["@odata.id"])
                            raids.append({
                                "RAIDType": vd.get("RAIDType", "N/A"),
                                "CapacityGB": round(vd.get("CapacityBytes", 0) / (1024 ** 3), 2),
                                "Health": vd.get("Status", {}).get("Health", "N/A"),
                            })
                    except InspectorError:
                        pass

                for drive in d.get("Drives", []):
                    try:
                        dd = self._get(drive["@odata.id"])
                        size = dd.get("CapacityBytes", 0)
                        disks.append({
                            "Model": dd.get("Name", "N/A"),
                            "State": dd.get("Oem", {}).get("Dell", {}).get("RaidStatus", "N/A"),
                            "Protocol": dd.get("Protocol", "N/A"),
                            "SizeGB": round(size / (1024 ** 3), 2) if size else "N/A",
                            "Health": dd.get("Status", {}).get("Health", "N/A"),
                        })
                    except InspectorError:
                        pass
        except InspectorError:
            pass
        result["storage_controllers"] = controllers
        result["raids"] = raids
        result["disks"] = disks

        # PSU
        psus = []
        psu_summary = {}
        try:
            pw = self._get("/redfish/v1/Chassis/System.Embedded.1/Power")
            sup = pw.get("PowerSupplies", [])
            psu_summary = {
                "Count": len(sup),
                "TotalCapacityW": sum(p.get("PowerCapacityWatts", 0) or 0 for p in sup),
                "Health": "OK" if all(p.get("Status", {}).get("Health") == "OK" for p in sup) else "Not OK",
            }
            redundancy = pw.get("Redundancy", [])
            if redundancy and isinstance(redundancy, list):
                psu_summary["Redundancy"] = redundancy[0].get("Status", {}).get("Health", "N/A")
            else:
                psu_summary["Redundancy"] = "N/A"
            for p in sup:
                psus.append({
                    "Name": p.get("Name", "N/A"),
                    "CapacityW": p.get("PowerCapacityWatts", "N/A"),
                    "Health": p.get("Status", {}).get("Health", "N/A"),
                })
        except InspectorError:
            pass
        result["psu_summary"] = psu_summary
        result["psus"] = psus

        # NIC (간단 정보)
        nics = []
        try:
            nl = self._get("/redfish/v1/Chassis/System.Embedded.1/NetworkAdapters")
            for m in nl.get("Members", []):
                ad = self._get(m["@odata.id"])
                model = ad.get("Model") or ad.get("Name", "N/A")
                aid = ad.get("Id", "N/A")
                for c in ad.get("Controllers", []):
                    ports = c.get("Links", {}).get("NetworkPorts", [])
                    for p in ports:
                        try:
                            pd = self._get(p["@odata.id"])
                            nics.append({
                                "Adapter": aid,
                                "Model": model,
                                "Port": pd.get("Id", "N/A"),
                                "LinkStatus": pd.get("LinkStatus", "N/A"),
                            })
                        except InspectorError:
                            pass
        except InspectorError:
            pass
        result["nics"] = nics

        return result

    # ---------- FW ----------
    def fetch_firmware(self) -> dict:
        """주의: 전체 컴포넌트는 항상 수집해두고, 필터/전체 표시는 포매터에서 선택"""
        result = {"service_tag": "N/A", "components": [], "all_components": [], "nic_fw": []}
        try:
            sys_data = self._get("/redfish/v1/Systems/System.Embedded.1")
            result["service_tag"] = sys_data.get("SKU", "N/A")
        except InspectorError:
            pass

        try:
            fw = self._get("/redfish/v1/UpdateService/FirmwareInventory")
            for m in fw.get("Members", []):
                mid = m.get("@odata.id", "")
                if "Installed" not in mid:
                    continue
                try:
                    c = self._get(mid)
                    name = c.get("Name", "")
                    version = c.get("Version", "N/A")
                    updateable = c.get("Updateable", "N/A")
                    entry = {
                        "Name": name,
                        "Version": version,
                        "Id": mid.rsplit("/", 1)[-1],
                        "Updateable": updateable,
                    }
                    # 핵심 컴포넌트 필터 (R6615/R740/R640 등 다양한 모델에서 운영자가 자주 확인하는 것들)
                    core_keywords = [
                        "BIOS",
                        "Lifecycle Controller",
                        "Integrated Dell Remote Access Controller",  # iDRAC
                        "PERC",
                        "BOSS",
                        "System CPLD",
                        "Power Supply",
                        "Backplane",
                        "TPM",
                        "Driver Pack",
                        "iDRAC Service Module",
                        "Diagnostics",
                    ]
                    if any(x in name for x in core_keywords):
                        result["components"].append({"Name": name, "Version": version})
                    # 전체는 항상 저장
                    result["all_components"].append(entry)
                except InspectorError:
                    continue
        except InspectorError:
            pass

        # NetworkAdapters/Members 에서 실제 존재하는 어댑터만 동적으로 (R6615/R740/R660/R760/R640 등 모델 무관)
        try:
            na = self._get("/redfish/v1/Chassis/System.Embedded.1/NetworkAdapters")
            for m in na.get("Members", []):
                slot_path = m.get("@odata.id", "")
                try:
                    d = self._get(slot_path)
                except InspectorError:
                    continue
                if not d:
                    continue
                slot_id = slot_path.rsplit("/", 1)[-1]
                name = d.get("Name", "N/A") if "Embedded" in slot_id else d.get("Model", d.get("Name", "N/A"))
                ver = "N/A"
                controllers = d.get("Controllers", [])
                if controllers:
                    ver = controllers[0].get("FirmwarePackageVersion", "N/A")
                result["nic_fw"].append({
                    "Slot": slot_id,
                    "Model": name,
                    "FirmwareVersion": ver,
                })
        except InspectorError:
            pass
        return result

    # ---------- BIOS ----------
    def fetch_bios(self) -> dict:
        try:
            bios = self._get("/redfish/v1/Systems/System.Embedded.1/Bios")
        except InspectorError as e:
            raise
        attrs = bios.get("Attributes", {})

        power_attrs = {}
        try:
            pw = self._get("/redfish/v1/Managers/System.Embedded.1/Attributes")
            power_attrs = pw.get("Attributes", {})
        except InspectorError:
            pass

        return {
            "system_profile": {
                "System Profile": attrs.get("SysProfile", "N/A"),
                "CPU Power Management": attrs.get("ProcPwrPerf", "N/A"),
                "C States": attrs.get("ProcCStates", "N/A"),
                "Memory Frequency": attrs.get("MemFrequency", "N/A"),
                "Turbo Boost": attrs.get("ProcTurboMode", "N/A"),
                "Memory Patrol Scrub": attrs.get("MemPatrolScrub", "N/A"),
                "Determinism Slider": attrs.get("DeterminismSlider", "N/A"),
                "Power Profile Select": attrs.get("PowerProfileSelect", "N/A"),
                "Algorithm Performance Boost": attrs.get("ApbDis", "N/A"),
                "DF C-State": attrs.get("DfCState", "N/A"),
            },
            "processor": {
                "Logical Processor": attrs.get("LogicalProc", "N/A"),
                "Virtualization Technology": attrs.get("ProcVirtualization", "N/A"),
                "NUMA Nodes Per Socket": attrs.get("NumaNodesPerSocket", "N/A"),
                "IOMMU Support": attrs.get("IommuSupport", "N/A"),
            },
            "integrated": {
                "SR-IOV Global Enable": attrs.get("SriovGlobalEnable", "N/A"),
            },
            "power_config": {
                "Redundancy Policy": power_attrs.get("ServerPwr.1.PSRedPolicy", "N/A"),
                "Hot Spare": power_attrs.get("ServerPwr.1.PSRapidOn", "N/A"),
            },
            # show_all 모드에서 사용할 전체 속성
            "all_attributes": dict(sorted(attrs.items())),
            "all_power_attributes": dict(sorted(power_attrs.items())),
        }


# =========================================================
#  Sample 데이터
# =========================================================
def sample_payload(kind: str) -> dict:
    if kind == "hw":
        return {
            "system": {"Model": "PowerEdge R760", "Manufacturer": "Dell Inc.", "ServiceTag": "SMPL123",
                       "PowerState": "On", "Health": "OK"},
            "cpu_summary": {"Model": "Intel(R) Xeon(R) Gold 6442Y", "Count": 2, "Health": "OK"},
            "cpus": [{"Socket": "CPU.Socket.1", "Speed": 2400, "Cores": 24},
                     {"Socket": "CPU.Socket.2", "Speed": 2400, "Cores": 24}],
            "memory_summary": {"TotalGiB": 256, "Health": "OK"},
            "memory_modules": [
                {"Location": "DIMM.Socket.A1", "CapacityGB": 32, "SpeedMHz": 4800},
                {"Location": "DIMM.Socket.A2", "CapacityGB": 32, "SpeedMHz": 4800},
                {"Location": "DIMM.Socket.B1", "CapacityGB": 32, "SpeedMHz": 4800},
                {"Location": "DIMM.Socket.B2", "CapacityGB": 32, "SpeedMHz": 4800},
            ],
            "fans": [{"Name": f"Fan {i+1}", "Health": "OK"} for i in range(6)],
            "storage_controllers": [{"Name": "PERC H965i Front", "Health": "OK"}],
            "raids": [{"RAIDType": "RAID1", "CapacityGB": 893.75, "Health": "OK"}],
            "disks": [
                {"Model": "Solid State Disk 0:1:0", "State": "Online", "Protocol": "SAS", "SizeGB": 893.75, "Health": "OK"},
                {"Model": "Solid State Disk 0:1:1", "State": "Online", "Protocol": "SAS", "SizeGB": 893.75, "Health": "OK"},
            ],
            "psu_summary": {"Count": 2, "TotalCapacityW": 2400, "Health": "OK", "Redundancy": "OK"},
            "psus": [{"Name": "PS1", "CapacityW": 1200, "Health": "OK"},
                     {"Name": "PS2", "CapacityW": 1200, "Health": "OK"}],
            "nics": [{"Adapter": "NIC.Integrated.1", "Model": "Broadcom 5720", "Port": "1", "LinkStatus": "Up"},
                     {"Adapter": "NIC.Integrated.1", "Model": "Broadcom 5720", "Port": "2", "LinkStatus": "Up"}],
        }
    if kind == "fw":
        all_components = [
            {"Name": "BIOS", "Version": "1.6.5", "Id": "BIOS.Setup.1-1", "Updateable": True},
            {"Name": "Lifecycle Controller", "Version": "6.10.80.00", "Id": "USC.Embedded.1:LC.Embedded.1", "Updateable": True},
            {"Name": "Integrated Dell Remote Access Controller", "Version": "7.00.30.00", "Id": "iDRAC.Embedded.1-1", "Updateable": True},
            {"Name": "System CPLD", "Version": "1.6.5", "Id": "CPLD.Embedded.1", "Updateable": False},
            {"Name": "PERC H965i Front", "Version": "8.6.0.0", "Id": "RAID.Embedded.1-1", "Updateable": True},
            {"Name": "BOSS-N1 Monolithic", "Version": "2.1.13.2025", "Id": "BOSS.SL.10-1", "Updateable": True},
            {"Name": "Power Supply.Slot.1", "Version": "00.18.31", "Id": "PSU.Slot.1", "Updateable": True},
            {"Name": "Power Supply.Slot.2", "Version": "00.18.31", "Id": "PSU.Slot.2", "Updateable": True},
            {"Name": "Backplane 1", "Version": "7.10", "Id": "Backplane.1", "Updateable": True},
            {"Name": "TPM", "Version": "7.2.3.1", "Id": "TPM.Integrated.1-1", "Updateable": False},
            {"Name": "Dell OS Driver Pack, 24.08.10, A00", "Version": "24.08.10", "Id": "DriverPack", "Updateable": True},
            {"Name": "Dell iDRAC Service Module Embedded Package v5.4.0.0", "Version": "5.4.0.0", "Id": "ServiceModule", "Updateable": True},
            {"Name": "Dell 64 Bit uEFI Diagnostics", "Version": "4303A24", "Id": "Diagnostics", "Updateable": True},
            {"Name": "Broadcom NetXtreme Gigabit Ethernet (BCM5720)", "Version": "22.91.5", "Id": "NIC.Embedded.1-1-1", "Updateable": True},
            {"Name": "Broadcom NetXtreme Gigabit Ethernet (BCM5720)", "Version": "22.91.5", "Id": "NIC.Embedded.2-1-1", "Updateable": True},
        ]
        return {
            "service_tag": "SMPL123",
            "components": [c for c in all_components if any(x in c["Name"] for x in
                ["BIOS","Lifecycle Controller","iDRAC","PERC","BOSS","CPLD","Power Supply","Backplane","TPM","Driver Pack","iDRAC Service Module","Diagnostics"])],
            "all_components": all_components,
            "nic_fw": [
                {"Slot": "NIC.Embedded.1", "Model": "Broadcom NetXtreme Gigabit Ethernet", "FirmwareVersion": "22.91.5"},
                {"Slot": "NIC.Embedded.2", "Model": "Broadcom NetXtreme Gigabit Ethernet", "FirmwareVersion": "22.91.5"},
            ],
        }
    if kind == "bios":
        # 전체 속성 샘플 — 그룹별로 다양한 prefix
        all_attrs = {
            # System Profile / Power
            "SysProfile": "PerfPerWattOptimizedOs", "ProcPwrPerf": "OsDbpm",
            "PowerSaver": "Disabled", "AcPwrRcvry": "Last", "AcPwrRcvryDelay": "Immediate",
            "ApbDis": "Disabled", "DfCState": "Enabled", "EnergyPerformanceBias": "BalancedPerformance",
            # Processor (AMD + Intel mix)
            "ProcCStates": "Enabled", "ProcTurboMode": "Enabled", "ProcVirtualization": "Enabled",
            "LogicalProc": "Enabled", "ProcAdjCacheLine": "Enabled", "ProcHwPrefetcher": "Enabled",
            "AgesaVersion": "1.0.0.7c", "BoostFMax": "Auto", "CcdCores": "All",
            "DcuIpPrefetcher": "Enabled", "DcuStreamerPrefetcher": "Enabled", "AvxIccpPregrant": "AllCoreOnly",
            # Memory
            "MemFrequency": "MaxPerf", "MemPatrolScrub": "Standard", "MemTest": "Disabled",
            "NumaNodesPerSocket": "1", "DimmSlot00": "Populated", "DimmSlot01": "Empty",
            "DramRefreshDelay": "Minimum", "AdddcSetting": "Disabled",
            # Boot
            "BootMode": "Uefi", "BootSeqRetry": "Enabled", "SecureBoot": "Enabled",
            "UefiVariableAccess": "Standard", "HddPlaceholder": "Disabled",
            # Network
            "PxeDev1EnDis": "Enabled", "PxeDev2EnDis": "Disabled", "HttpDev1EnDis": "Disabled",
            "IscsiDev1EnDis": "Disabled",
            # PCIe / Slot
            "SriovGlobalEnable": "Disabled", "Slot1": "Enabled", "Slot2": "Enabled",
            "EmbeddedSata": "AhciMode", "EmbeddedVideo": "Enabled", "IntegratedNetwork1": "Enabled",
            # Security
            "TpmSecurity": "On", "TpmFirmware": "7.2.3.1", "AesNi": "Enabled",
            "AuthorizeDeviceFirmware": "Disabled", "SignedFirmwareUpdate": "Enabled",
            # USB
            "UsbPorts": "AllOn", "InternalUsb": "On", "FrontUsb": "On", "OsWatchdogTimer": "Disabled",
            # System Info
            "AssetTag": "RACK-A-12", "ServiceTag": "SMPL123", "SystemBiosMessage": "Disabled",
            "IommuSupport": "Enabled",
        }
        return {
            "system_profile": {
                "System Profile": "PerfPerWattOptimizedOs", "CPU Power Management": "OsDbpm",
                "C States": "Enabled", "Memory Frequency": "MaxPerf", "Turbo Boost": "Enabled",
                "Memory Patrol Scrub": "Standard", "Determinism Slider": "N/A",
                "Power Profile Select": "N/A", "Algorithm Performance Boost": "N/A", "DF C-State": "N/A",
            },
            "processor": {"Logical Processor": "Enabled", "Virtualization Technology": "Enabled",
                          "NUMA Nodes Per Socket": "1", "IOMMU Support": "Enabled"},
            "integrated": {"SR-IOV Global Enable": "Disabled"},
            "power_config": {"Redundancy Policy": "Redundant", "Hot Spare": "Disabled"},
            "all_attributes": dict(sorted(all_attrs.items())),
            "all_power_attributes": {
                "ServerPwr.1.PSRedPolicy": "Redundant",
                "ServerPwr.1.PSRapidOn": "Disabled",
                "iDRAC.NIC.1.Enable": "Enabled",
                "iDRAC.NIC.1.MTU": "1500",
            },
        }
    return {}


# =========================================================
#  포매터 (HTML + plain text + xlsx 행)
# =========================================================
def html_section(title: str) -> str:
    """소섹션 바 — Qt QTextBrowser에서 확실하게 색상이 보이도록 테이블 사용.
       엑셀 디자인의 연파랑(#D5E5F2) + 진파랑 텍스트(#0E639C) + 왼쪽 진파랑 보더"""
    return (
        '<br>'
        '<table cellspacing="0" cellpadding="0" width="100%" '
        'style="border-collapse:collapse;">'
        '<tr>'
        '<td bgcolor="#0E639C" width="5" style="background-color:#0E639C;">&nbsp;</td>'
        '<td bgcolor="#D5E5F2" '
        'style="background-color:#D5E5F2; padding:8px 14px;">'
        f'<span style="color:#0E639C; font-weight:700; font-size:14px;">'
        f'{html.escape(title)}</span>'
        '</td>'
        '</tr>'
        '</table>'
    )


def html_main_section(title: str, color: str = "#264F78") -> str:
    """큰 섹션 바 (HW/FW/BIOS) — 엑셀의 네이비 바와 동일 톤. 테이블 셀 사용"""
    return (
        '<br><br>'
        '<table cellspacing="0" cellpadding="0" width="100%" '
        'style="border-collapse:collapse;">'
        '<tr>'
        f'<td bgcolor="{color}" '
        f'style="background-color:{color}; padding:11px 18px;">'
        f'<span style="color:#FFFFFF; font-weight:700; font-size:16px;">'
        f'{html.escape(title)}</span>'
        '</td>'
        '</tr>'
        '</table>'
    )


def html_divider() -> str:
    """큰 섹션 간 구분선"""
    return '<br><hr style="border:0; border-top:1px solid #d0d7de; margin:24px 0;"><br>'


def html_kv_table(rows: List[Tuple[str, str]], allow_html_value: bool = True) -> str:
    """key는 항상 escape, value는 기본적으로 raw HTML 허용(health_html 등 사용 위해).
    사용자 입력값을 직접 넣을 때는 미리 html.escape() 처리해서 전달할 것."""
    body = ""
    for k, v in rows:
        v_str = str(v) if v is not None else ""
        if not allow_html_value:
            v_str = html.escape(v_str)
        body += (
            f'<tr>'
            f'<td style="padding:5px 16px 5px 6px; color:#57606a; font-size:12px;">{html.escape(str(k))}</td>'
            f'<td style="padding:5px 6px; color:#1f2328;"><b>{v_str}</b></td>'
            f'</tr>'
        )
    return f'<table cellspacing="0" style="border-collapse:collapse;">{body}</table>'


def health_html(value: str) -> str:
    v = str(value)
    color = "#999"
    if v == "OK":
        color = "#1a7f37"
    elif v in ("Critical", "Not OK", "Failed"):
        color = "#d1242f"
    elif v == "Warning":
        color = "#bf8700"
    return f'<span style="color:{color}; font-weight:600;">{html.escape(v)}</span>'


def _e(v) -> str:
    """값을 안전하게 HTML로 만들기 (str escape)"""
    return html.escape(str(v)) if v is not None else ""


def format_hw_html(data: dict) -> str:
    s = data.get("system", {})
    out = html_section("System Info")
    out += html_kv_table([
        ("모델명", _e(s.get("Model"))), ("제조사", _e(s.get("Manufacturer"))),
        ("Service Tag", _e(s.get("ServiceTag"))), ("전원 상태", _e(s.get("PowerState"))),
        ("시스템 상태", health_html(s.get("Health", "N/A"))),
    ])

    cs = data.get("cpu_summary", {})
    out += html_section("Processors")
    out += html_kv_table([
        ("CPU 모델", _e(cs.get("Model"))), ("수량", _e(cs.get("Count"))),
        ("상태", health_html(cs.get("Health", "N/A"))),
    ])
    if data.get("cpus"):
        out += '<div style="margin:6px 0 0 8px; color:#444;">'
        for c in data["cpus"]:
            out += f'• Socket {html.escape(str(c["Socket"]))}: {c["Speed"]} MHz, {c["Cores"]} cores<br>'
        out += '</div>'

    ms = data.get("memory_summary", {})
    out += html_section("System Memory")
    out += html_kv_table([
        ("총 용량", _e(f"{ms.get('TotalGiB', 'N/A')} GiB")),
        ("상태", health_html(ms.get("Health", "N/A"))),
        ("DIMM 수량", _e(len(data.get("memory_modules", [])))),
    ])
    if data.get("memory_modules"):
        out += '<div style="margin:6px 0 0 8px; color:#444;">'
        for m in data["memory_modules"]:
            out += f'• {html.escape(str(m["Location"]))}: {m["CapacityGB"]} GB @ {m["SpeedMHz"]} MHz<br>'
        out += '</div>'

    out += html_section("FAN")
    fans = data.get("fans", [])
    out += html_kv_table([
        ("수량", _e(len(fans))),
        ("전체 상태",
         health_html("OK" if all(f["Health"] == "OK" for f in fans) else "Not OK") if fans else _e("N/A")),
    ])

    out += html_section("Storage Controller")
    if data.get("storage_controllers"):
        out += '<div style="margin:0 0 0 8px; color:#444;">'
        for c in data["storage_controllers"]:
            out += f'• {html.escape(c["Name"])} ({health_html(c["Health"])})<br>'
        out += '</div>'
    else:
        out += '<div style="color:#999; margin-left:8px;">(컨트롤러 없음)</div>'

    out += html_section("RAID Configuration")
    if data.get("raids"):
        out += '<div style="margin:0 0 0 8px; color:#444;">'
        for r in data["raids"]:
            out += f'• {r["RAIDType"]}, {r["CapacityGB"]} GB, {health_html(r["Health"])}<br>'
        out += '</div>'
    else:
        out += '<div style="color:#999; margin-left:8px;">(RAID 없음)</div>'

    out += html_section("Physical Disks")
    if data.get("disks"):
        out += '<div style="margin:0 0 0 8px; color:#444;">'
        for d in data["disks"]:
            out += (f'• {html.escape(d["Model"])} — {d["State"]}, {d["Protocol"]}, '
                    f'{d["SizeGB"]} GB ({health_html(d["Health"])})<br>')
        out += '</div>'
    else:
        out += '<div style="color:#999; margin-left:8px;">(디스크 정보 없음)</div>'

    ps = data.get("psu_summary", {})
    out += html_section("Power Supply")
    out += html_kv_table([
        ("PSU 수량", _e(ps.get("Count"))),
        ("총 전력 용량", _e(f"{ps.get('TotalCapacityW', 'N/A')} W")),
        ("상태", health_html(ps.get("Health", "N/A"))),
        ("이중화", health_html(ps.get("Redundancy", "N/A"))),
    ])

    out += html_section("Network Adapters")
    if data.get("nics"):
        out += '<div style="margin:0 0 0 8px; color:#444;">'
        for n in data["nics"]:
            out += (f'• {html.escape(n["Model"])} ({html.escape(n["Adapter"])}) '
                    f'Port {n["Port"]} — {health_html(n["LinkStatus"])}<br>')
        out += '</div>'
    else:
        out += '<div style="color:#999; margin-left:8px;">(어댑터 없음)</div>'
    return out


# --- BIOS 속성 그룹핑 (show_all 시 사용) ---
# R6615(AMD), R740(Intel), R6515/R6525/R650/R660/R750/R760 등 다양한 모델 대응
BIOS_GROUPS = [
    ("System Profile / Power",  ["SysProfile", "Power", "Determinism", "Apb", "Df", "DfP", "Energy", "AcPwr", "ProcPwr"]),
    ("Processor (AMD)",         ["Agesa", "Ccx", "Ccd", "Boost", "BoostFMax", "DfCState", "DfPstate", "ControlledTurbo", "Cpu", "CpuAcpi", "CpuMinSev", "CpuInter"]),
    ("Processor (Intel)",       ["Proc", "Logical", "Hyper", "Dcu", "Avx", "Direct", "Sva", "Llc", "Snc", "Adddc"]),
    ("Memory",                  ["Mem", "Dimm", "Patrol", "Node", "Numa", "Dram", "Cxl", "OppSrefEn"]),
    ("Boot / UEFI",             ["Boot", "Uefi", "Set", "OneTime", "Hdd", "BiosBoot", "GenericUsb"]),
    ("Network / PXE / HTTP / iSCSI", ["Pxe", "Http", "Network", "Iscsi"]),
    ("PCIe / Slot / SR-IOV",    ["Slot", "Pci", "Sriov", "Embedded", "Integrated"]),
    ("Storage / SATA / NVMe",   ["Sata", "Nvme", "Storage", "Boss", "EmbSata", "EmbVideo"]),
    ("USB / Front Panel",       ["Usb", "Internal", "Front", "Os", "GenericUsb"]),
    ("Security / TPM",          ["Tpm", "Sec", "Aes", "MemEnc", "SignedFirmware", "Inband", "Authorize"]),
    ("Redundant OS / Watchdog", ["RedundantOs", "Watch", "ErrPrompt", "Err", "CECrit"]),
    ("Date / Time / Console",   ["Date", "Time", "Daylight", "ConTerm"]),
    ("System Information",      ["SystemBiosMessage", "Service", "Asset", "Dell", "AgesaVersion", "Current"]),
]

# 5g 모드 (기본 핵심 필터) — 운영자가 가장 자주 확인하는 항목만
PRESET_5G_KEYS = {
    # System Profile
    "SysProfile", "ProcPwrPerf", "ProcCStates", "ProcTurboMode",
    "MemFrequency", "MemPatrolScrub",
    # Processor
    "LogicalProc", "ProcVirtualization", "NumaNodesPerSocket", "IommuSupport",
    # Integrated
    "SriovGlobalEnable",
    # Boot
    "BootMode", "SecureBoot",
    # Network/PXE
    "PxeDev1EnDis", "PxeDev2EnDis",
}


def parse_key_filter(filter_text: str) -> List[str]:
    """콤마/공백 구분 키워드 파싱. 빈 문자열이면 빈 리스트 반환 (= 필터 OFF)."""
    if not filter_text:
        return []
    raw = filter_text.replace("\n", ",").replace(" ", ",").split(",")
    return [t.strip() for t in raw if t.strip()]


def match_key(key: str, patterns: List[str]) -> bool:
    """key가 어느 하나의 패턴이라도 매칭되면 True. 와일드카드 * 지원, 대소문자 무시"""
    if not patterns:
        return True
    import fnmatch
    kl = key.lower()
    for p in patterns:
        pl = p.lower()
        if "*" in pl or "?" in pl:
            if fnmatch.fnmatch(kl, pl):
                return True
        else:
            if pl in kl:  # 부분 일치
                return True
    return False


def group_bios_attrs(attrs: dict) -> List[Tuple[str, List[Tuple[str, str]]]]:
    """BIOS 속성을 prefix 그룹별로 묶기"""
    keys = sorted(attrs.keys())
    grouped: dict = {name: [] for name, _ in BIOS_GROUPS}
    grouped["Misc / 기타"] = []
    used = set()
    for k in keys:
        placed = False
        for name, prefixes in BIOS_GROUPS:
            if any(k.startswith(p) for p in prefixes):
                grouped[name].append((k, str(attrs[k])))
                placed = True
                break
        if not placed:
            grouped["Misc / 기타"].append((k, str(attrs[k])))
    return [(n, items) for n, items in grouped.items() if items]


def format_fw_html(data: dict, show_all: bool = False, key_filter: Optional[List[str]] = None) -> str:
    out = html_section("Service Tag")
    out += html_kv_table([("Service Tag", _e(data.get("service_tag", "N/A")))])

    if show_all:
        comps_all = data.get("all_components", [])
        if key_filter:
            comps = [c for c in comps_all if match_key(c.get("Name", "") + " " + c.get("Id", ""), key_filter)]
            out += html_section(f"Firmware — 필터 매칭 ({len(comps)}/{len(comps_all)}개)")
        else:
            comps = comps_all
            out += html_section(f"Firmware — 전체 컴포넌트 ({len(comps)}개)")
        if comps:
            out += '<div style="margin:0 0 0 8px; color:#444; line-height:1.7;">'
            for c in comps:
                out += f'• {html.escape(c["Name"])} — <b>{html.escape(c["Version"])}</b><br>'
            out += '</div>'
        else:
            out += '<div style="color:#999; margin-left:8px;">(매칭되는 컴포넌트 없음)</div>'
    else:
        out += html_section("Firmware (핵심 컴포넌트)")
        comps = data.get("components", [])
        if key_filter:
            comps = [c for c in comps if match_key(c.get("Name", ""), key_filter)]
        if comps:
            out += '<div style="margin:0 0 0 8px; color:#444;">'
            for c in comps:
                out += f'• {html.escape(c["Name"])} — <b>{html.escape(c["Version"])}</b><br>'
            out += '</div>'
        else:
            out += '<div style="color:#999; margin-left:8px;">(매칭되는 컴포넌트 없음)</div>'

    out += html_section("NIC Firmware")
    if data.get("nic_fw"):
        out += '<div style="margin:0 0 0 8px; color:#444;">'
        for n in data["nic_fw"]:
            out += f'• {html.escape(n["Slot"])} / {html.escape(n["Model"])} — <b>{html.escape(n["FirmwareVersion"])}</b><br>'
        out += '</div>'
    else:
        out += '<div style="color:#999; margin-left:8px;">(NIC 정보 없음)</div>'
    return out


def format_bios_html(data: dict, show_all: bool = False, preset_5g: bool = False,
                     key_filter: Optional[List[str]] = None) -> str:
    out = ""
    # 키 필터가 있으면 모드와 무관하게 전체 속성에서 매칭만 표시
    def _esc_items(pairs):
        return [(k, _e(v)) for k, v in pairs]

    if key_filter:
        attrs = data.get("all_attributes", {})
        items = [(k, _e(attrs[k])) for k in sorted(attrs.keys()) if match_key(k, key_filter)]
        out += html_section(f"BIOS — 키 필터 매칭 ({len(items)}/{len(attrs)}개)")
        if items:
            out += html_kv_table(items)
        else:
            out += '<div style="color:#999; margin-left:8px;">(매칭되는 속성 없음)</div>'
        pwr = data.get("all_power_attributes", {})
        pitems = [(k, _e(v)) for k, v in pwr.items() if match_key(k, key_filter)]
        if pitems:
            out += html_section(f"iDRAC Manager Attributes — 매칭 ({len(pitems)}개)")
            out += html_kv_table(pitems)
        return out

    if show_all:
        attrs = data.get("all_attributes", {})
        out += html_section(f"BIOS 전체 속성 ({len(attrs)}개)")
        for name, items in group_bios_attrs(attrs):
            out += html_section(name)
            out += html_kv_table(_esc_items(items))
        pwr = data.get("all_power_attributes", {})
        if pwr:
            out += html_section(f"iDRAC Manager Attributes ({len(pwr)}개)")
            out += html_kv_table([(k, _e(v)) for k, v in pwr.items()])
        return out

    if preset_5g:
        attrs = data.get("all_attributes", {})
        items = [(k, _e(attrs[k])) for k in sorted(attrs.keys()) if k in PRESET_5G_KEYS]
        out += html_section(f"BIOS — 5g 모드 (운영 핵심 {len(items)}개)")
        if items:
            out += html_kv_table(items)
        else:
            out += '<div style="color:#999; margin-left:8px;">(해당 속성 없음)</div>'
        return out

    # 기본 (기존 필터 뷰)
    out += html_section("System Profile Settings")
    out += html_kv_table(_esc_items(list(data.get("system_profile", {}).items())))
    out += html_section("Processor Settings")
    out += html_kv_table(_esc_items(list(data.get("processor", {}).items())))
    out += html_section("Integrated Devices")
    out += html_kv_table(_esc_items(list(data.get("integrated", {}).items())))
    out += html_section("Power Configuration")
    out += html_kv_table(_esc_items(list(data.get("power_config", {}).items())))
    return out


# Plain text 포매터 (보기 좋게 OFF + TXT 저장용)
def format_hw_text(data: dict) -> str:
    L = []
    sep = "=" * 60
    def section(t):
        L.append(""); L.append(sep); L.append(t.center(60)); L.append(sep)

    s = data.get("system", {})
    section("System Info")
    L += [f"  모델명       : {s.get('Model')}",
          f"  제조사       : {s.get('Manufacturer')}",
          f"  Service Tag : {s.get('ServiceTag')}",
          f"  전원 상태    : {s.get('PowerState')}",
          f"  시스템 상태  : {s.get('Health')}"]

    cs = data.get("cpu_summary", {})
    section("Processors")
    L += [f"  CPU 모델 : {cs.get('Model')}",
          f"  수량     : {cs.get('Count')}",
          f"  상태     : {cs.get('Health')}"]
    for c in data.get("cpus", []):
        L.append(f"    - Socket {c['Socket']}: {c['Speed']} MHz, {c['Cores']} cores")

    ms = data.get("memory_summary", {})
    section("System Memory")
    L += [f"  총 용량   : {ms.get('TotalGiB')} GiB",
          f"  상태      : {ms.get('Health')}",
          f"  DIMM 수량 : {len(data.get('memory_modules', []))}"]
    for m in data.get("memory_modules", []):
        L.append(f"    - {m['Location']}: {m['CapacityGB']} GB @ {m['SpeedMHz']} MHz")

    section("FAN")
    fans = data.get("fans", [])
    L += [f"  수량      : {len(fans)}",
          f"  전체 상태 : {'OK' if fans and all(f['Health']=='OK' for f in fans) else ('Not OK' if fans else 'N/A')}"]

    section("Storage Controllers")
    for c in data.get("storage_controllers", []):
        L.append(f"  - {c['Name']} ({c['Health']})")

    section("RAID Configuration")
    for r in data.get("raids", []):
        L.append(f"  - {r['RAIDType']}, {r['CapacityGB']} GB, {r['Health']}")

    section("Physical Disks")
    for d in data.get("disks", []):
        L.append(f"  - {d['Model']} | {d['State']} | {d['Protocol']} | {d['SizeGB']} GB | {d['Health']}")

    ps = data.get("psu_summary", {})
    section("Power Supply")
    L += [f"  PSU 수량        : {ps.get('Count')}",
          f"  총 전력 용량    : {ps.get('TotalCapacityW')} W",
          f"  전체 상태       : {ps.get('Health')}",
          f"  이중화          : {ps.get('Redundancy')}"]

    section("Network Adapters")
    for n in data.get("nics", []):
        L.append(f"  - {n['Model']} ({n['Adapter']}) Port {n['Port']} — {n['LinkStatus']}")

    return "\n".join(L)


def format_fw_text(data: dict, show_all: bool = False, key_filter: Optional[List[str]] = None) -> str:
    L = []; sep = "=" * 60
    L += ["", sep, "Service TAG".center(60), sep, f"Service Tag: {data.get('service_tag')}"]
    if show_all:
        comps = data.get("all_components", [])
        if key_filter:
            comps = [c for c in comps if match_key(c.get("Name", "") + " " + c.get("Id", ""), key_filter)]
            L += ["", sep, f"Firmware - 필터 매칭 ({len(comps)}개)".center(60), sep]
        else:
            L += ["", sep, f"Firmware - 전체 ({len(comps)}개)".center(60), sep]
        for c in comps:
            L.append(f" {c['Name'][:38]:38} {c['Version']}")
    else:
        comps = data.get("components", [])
        if key_filter:
            comps = [c for c in comps if match_key(c.get("Name", ""), key_filter)]
        L += ["", sep, "Firmware".center(60), sep]
        for c in comps:
            L.append(f" {c['Name']:30} {c['Version']}")
    L += ["", sep, "NIC Firmware".center(60), sep]
    for n in data.get("nic_fw", []):
        L.append(f" {n['Slot']:25} {n['Model'][:25]:25} {n['FirmwareVersion']}")
    return "\n".join(L)


def format_bios_text(data: dict, show_all: bool = False, preset_5g: bool = False,
                     key_filter: Optional[List[str]] = None) -> str:
    L = []; sep = "=" * 60
    if key_filter:
        attrs = data.get("all_attributes", {})
        items = [(k, attrs[k]) for k in sorted(attrs.keys()) if match_key(k, key_filter)]
        L += ["", sep, f"BIOS - 키 필터 매칭 ({len(items)}/{len(attrs)}개)".center(60), sep]
        for k, v in items:
            L.append(f" {k:38} {v}")
        pwr = data.get("all_power_attributes", {})
        pitems = [(k, v) for k, v in pwr.items() if match_key(k, key_filter)]
        if pitems:
            L += ["", "-" * 60, f" iDRAC Manager 매칭 ({len(pitems)}개)", "-" * 60]
            for k, v in pitems:
                L.append(f" {k:38} {v}")
        return "\n".join(L)
    if show_all:
        attrs = data.get("all_attributes", {})
        L += ["", sep, f"BIOS 전체 속성 ({len(attrs)}개)".center(60), sep]
        for name, items in group_bios_attrs(attrs):
            L += ["", "-" * 60, f" {name}", "-" * 60]
            for k, v in items:
                L.append(f" {k:38} {v}")
        pwr = data.get("all_power_attributes", {})
        if pwr:
            L += ["", "-" * 60, f" iDRAC Manager Attributes ({len(pwr)}개)", "-" * 60]
            for k, v in pwr.items():
                L.append(f" {k:38} {v}")
        return "\n".join(L)

    if preset_5g:
        attrs = data.get("all_attributes", {})
        items = [(k, attrs[k]) for k in sorted(attrs.keys()) if k in PRESET_5G_KEYS]
        L += ["", sep, f"BIOS - 5g 모드 (운영 핵심 {len(items)}개)".center(60), sep]
        for k, v in items:
            L.append(f" {k:38} {v}")
        return "\n".join(L)

    for title, key in [("System Profile Settings", "system_profile"),
                       ("Processor Settings", "processor"),
                       ("Integrated Devices", "integrated"),
                       ("Power Configuration", "power_config")]:
        L += ["", sep, title.center(60), sep]
        for k, v in data.get(key, {}).items():
            L.append(f" {k:34} {v}")
    return "\n".join(L)


# =========================================================
#  Excel 누적 저장 (openpyxl)
# =========================================================
# 원본 PySide6 앱과 동일한 컬러 팔레트
XLSX_COLOR_MAIN_BG     = "0E639C"  # 메인 헤더 진파랑
XLSX_COLOR_SECTION_BG  = "264F78"  # 큰 섹션 네이비
XLSX_COLOR_SUBSEC_BG   = "D5E5F2"  # 소섹션 연파랑
XLSX_COLOR_META_BG     = "EEEEEE"  # 메타 라벨 회색
XLSX_COLOR_WHITE       = "FFFFFF"
XLSX_COLOR_BLUE_TEXT   = "0E639C"
XLSX_COLOR_DARK_TEXT   = "333333"


def save_combined_xlsx(path: str, service_tag: str, payloads: dict,
                       ip: str = "", options_str: str = "", timestamp: str = "") -> str:
    """HW + FW + BIOS 를 한 시트에 통합 저장.
       payloads = {'hw': {...}, 'fw': {...}, 'bios': {...}} 중 있는 것만 차례로.
       시트명: ServiceTag (중복 시 _2, _3 ...)
       반환: 사용된 시트 이름.
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    # 시트명: 그냥 ServiceTag, 충돌시 _2, _3 ...
    base = (service_tag or "Server")[:28]
    name = base
    i = 2
    while name in wb.sheetnames:
        name = f"{base}_{i}"[:31]
        i += 1
    ws = wb.create_sheet(name)

    # 컬럼 너비
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # 스타일 정의
    F_MAIN     = Font(bold=True, color=XLSX_COLOR_WHITE, size=12)
    F_SECTION  = Font(bold=True, color=XLSX_COLOR_WHITE, size=12)
    F_SUBSEC   = Font(bold=True, color=XLSX_COLOR_BLUE_TEXT, size=11)
    F_LABEL    = Font(bold=True, color=XLSX_COLOR_DARK_TEXT)
    F_VAL      = Font(color=XLSX_COLOR_DARK_TEXT)
    F_BULLET   = Font(color=XLSX_COLOR_DARK_TEXT)
    P_MAIN     = PatternFill("solid", fgColor=XLSX_COLOR_MAIN_BG)
    P_SECTION  = PatternFill("solid", fgColor=XLSX_COLOR_SECTION_BG)
    P_SUBSEC   = PatternFill("solid", fgColor=XLSX_COLOR_SUBSEC_BG)
    P_META     = PatternFill("solid", fgColor=XLSX_COLOR_META_BG)
    A_LEFT     = Alignment(horizontal="left", vertical="center", indent=1)
    A_LEFT_PLAIN = Alignment(horizontal="left", vertical="center")

    state = {"r": 1}

    def write_main_header(title):
        r = state["r"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=title)
        c.font = F_MAIN; c.fill = P_MAIN; c.alignment = A_LEFT
        ws.row_dimensions[r].height = 24
        state["r"] = r + 1

    def write_meta(label, value):
        r = state["r"]
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = F_LABEL; c1.fill = P_META; c1.alignment = A_LEFT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c2 = ws.cell(row=r, column=2, value=value)
        c2.font = F_VAL; c2.alignment = A_LEFT_PLAIN
        state["r"] = r + 1

    def write_section_bar(title):
        state["r"] += 2
        r = state["r"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=title)
        c.font = F_SECTION; c.fill = P_SECTION; c.alignment = A_LEFT
        ws.row_dimensions[r].height = 22
        state["r"] = r + 2

    def write_subsection_bar(title):
        state["r"] += 1
        r = state["r"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=title)
        c.font = F_SUBSEC; c.fill = P_SUBSEC; c.alignment = A_LEFT
        ws.row_dimensions[r].height = 20
        state["r"] = r + 1

    def write_kv(label, value):
        r = state["r"]
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = F_LABEL; c1.alignment = A_LEFT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c2 = ws.cell(row=r, column=2, value="" if value is None else str(value))
        c2.font = F_VAL; c2.alignment = A_LEFT_PLAIN
        state["r"] = r + 1

    def write_bullet(text):
        r = state["r"]
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=2, value=f"• {text}")
        c.font = F_BULLET; c.alignment = A_LEFT_PLAIN
        state["r"] = r + 1

    def write_inline_label(label):
        r = state["r"]
        c = ws.cell(row=r, column=1, value=f"  {label}")
        c.font = F_VAL; c.alignment = A_LEFT_PLAIN
        state["r"] = r + 1

    def gap():
        state["r"] += 1

    # ----- 메인 헤더 + 메타 4행 -----
    write_main_header("iDRAC 조회 결과")
    write_meta("Service Tag", service_tag or "N/A")
    write_meta("iDRAC IP", ip or "")
    write_meta("조회 옵션", options_str or "")
    write_meta("조회 시간", timestamp or "")

    # ===== HW =====
    if "hw" in payloads:
        hw = payloads["hw"]
        write_section_bar("Hardware Status && Health Check")

        write_subsection_bar("System info"); gap()
        s = hw.get("system", {})
        write_kv("모델명", s.get("Model"))
        write_kv("제조사", s.get("Manufacturer"))
        write_kv("서비스 태그", s.get("ServiceTag"))
        write_kv("전원 상태", s.get("PowerState"))
        write_kv("시스템 상태", s.get("Health"))

        write_subsection_bar("Processors"); gap()
        cs = hw.get("cpu_summary", {})
        write_kv("CPU 모델명", cs.get("Model"))
        write_kv("CPU 수량", cs.get("Count"))
        write_kv("CPU 상태", cs.get("Health"))
        write_inline_label("CPU 상세 정보:")
        for c in hw.get("cpus", []):
            write_bullet(f"Socket {c['Socket']}: {c['Speed']} MHz, {c['Cores']} cores")

        write_subsection_bar("System Memory"); gap()
        ms = hw.get("memory_summary", {})
        write_kv("Memory 총 용량", f"{ms.get('TotalGiB', 'N/A')} GiB")
        write_kv("Memory 수량", len(hw.get("memory_modules", [])))
        write_kv("Memory 상태", ms.get("Health"))
        write_inline_label("Memory 용량    :")
        for m in hw.get("memory_modules", []):
            write_bullet(f"{m['Location']}: {m['CapacityGB']} GB @ {m['SpeedMHz']} MHz")

        write_subsection_bar("FAN"); gap()
        fans = hw.get("fans", [])
        write_kv("FAN 수량", len(fans))
        write_kv("FAN 상태",
                 "OK" if fans and all(f.get("Health") == "OK" for f in fans) else
                 ("Not OK" if fans else "N/A"))

        write_subsection_bar("Storage Controller"); gap()
        for c in hw.get("storage_controllers", []):
            write_bullet(f"{c['Name']} (상태: {c['Health']})")

        write_subsection_bar("RAID Configuration"); gap()
        for r2 in hw.get("raids", []):
            write_bullet(f"RAID Level: {r2['RAIDType']}, 용량: {r2['CapacityGB']} GB, 상태: {r2['Health']}")

        write_subsection_bar("Physical Disks"); gap()
        for d in hw.get("disks", []):
            write_bullet(
                f"Model: {d['Model']}, State: {d['State']}, "
                f"Bus Protocol: {d['Protocol']}, Size: {d['SizeGB']} GB, Status: {d['Health']}"
            )

        write_subsection_bar("Power Supply"); gap()
        ps = hw.get("psu_summary", {})
        write_kv("PSU 수량", ps.get("Count"))
        write_kv("PSU 상태", ps.get("Health"))
        write_kv("PSU 총 전력 용량", f"{ps.get('TotalCapacityW', 'N/A')} W")
        write_kv("PSU 이중화 상태", ps.get("Redundancy"))

        write_subsection_bar("Network Adapters"); gap()
        for n in hw.get("nics", []):
            write_bullet(
                f"Model: {n['Model']}, Adapter: {n['Adapter']}, "
                f"Port: {n['Port']}, LinkStatus: {n['LinkStatus']}"
            )

    # ===== FW =====
    if "fw" in payloads:
        fw = payloads["fw"]
        write_section_bar("Firmware Information")

        write_subsection_bar("Service Tag"); gap()
        write_kv("Service Tag", fw.get("service_tag", "N/A"))

        use_all = (fw.get("all_components") and
                   len(fw["all_components"]) > len(fw.get("components", [])))

        if use_all:
            write_subsection_bar(f"Firmware - 전체 컴포넌트 ({len(fw['all_components'])}개)")
            gap()
            for c in fw["all_components"]:
                write_bullet(f"{c['Name']} — {c['Version']}")
        else:
            write_subsection_bar("Firmware (BIOS / LC / iDRAC / PERC / BOSS / CPLD / PSU / TPM 등)")
            gap()
            for c in fw.get("components", []):
                write_bullet(f"{c['Name']} — {c['Version']}")

        write_subsection_bar("NIC Firmware"); gap()
        for n in fw.get("nic_fw", []):
            write_bullet(f"{n['Slot']} / {n['Model']} — {n['FirmwareVersion']}")

    # ===== BIOS =====
    if "bios" in payloads:
        bios = payloads["bios"]
        write_section_bar("BIOS Configuration")

        attrs = bios.get("all_attributes") or {}
        if attrs and len(attrs) > 20:
            # show_all 모드: 전체 속성 그룹별로
            for name_grp, items in group_bios_attrs(attrs):
                write_subsection_bar(name_grp); gap()
                for k, v in items:
                    write_kv(k, v)
            pwr = bios.get("all_power_attributes") or {}
            if pwr:
                write_subsection_bar(f"iDRAC Manager Attributes ({len(pwr)}개)"); gap()
                for k, v in pwr.items():
                    write_kv(k, v)
        else:
            # 기본 (필터) 모드
            write_subsection_bar("System Setup → System BIOS → System Profile Settings"); gap()
            for k, v in bios.get("system_profile", {}).items():
                write_kv(k, v)
            write_subsection_bar("System Setup → System BIOS → Processor Settings"); gap()
            for k, v in bios.get("processor", {}).items():
                write_kv(k, v)
            write_subsection_bar("System Setup → System BIOS → Integrated Devices"); gap()
            for k, v in bios.get("integrated", {}).items():
                write_kv(k, v)
            write_subsection_bar("System Setup → Power Configuration"); gap()
            for k, v in bios.get("power_config", {}).items():
                write_kv(k, v)

    wb.save(path)
    return name


def save_to_xlsx(path: str, sheet_prefix: str, kind: str, payload: dict,
                 ip: str = "", options_str: str = "", timestamp: str = "") -> str:
    """원본 디자인 그대로 — 진파랑 헤더 / 네이비 큰섹션 / 연파랑 소섹션 / 불릿 리스트.
       기존 파일이 있으면 시트 추가, 없으면 새로 생성. 반환: 사용된 시트 이름."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    if os.path.exists(path):
        wb = load_workbook(path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    base = f"{sheet_prefix}_{kind.upper()}"[:28]
    name = base
    i = 2
    while name in wb.sheetnames:
        name = f"{base}_{i}"[:31]
        i += 1
    ws = wb.create_sheet(name)

    # 컬럼 너비 (원본과 동일)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # ---------- 스타일 헬퍼 ----------
    F_MAIN     = Font(bold=True, color=XLSX_COLOR_WHITE, size=12)
    F_SECTION  = Font(bold=True, color=XLSX_COLOR_WHITE, size=12)
    F_SUBSEC   = Font(bold=True, color=XLSX_COLOR_BLUE_TEXT, size=11)
    F_LABEL    = Font(bold=True, color=XLSX_COLOR_DARK_TEXT)
    F_VAL      = Font(color=XLSX_COLOR_DARK_TEXT)
    F_BULLET   = Font(color=XLSX_COLOR_DARK_TEXT)
    P_MAIN     = PatternFill("solid", fgColor=XLSX_COLOR_MAIN_BG)
    P_SECTION  = PatternFill("solid", fgColor=XLSX_COLOR_SECTION_BG)
    P_SUBSEC   = PatternFill("solid", fgColor=XLSX_COLOR_SUBSEC_BG)
    P_META     = PatternFill("solid", fgColor=XLSX_COLOR_META_BG)
    A_LEFT     = Alignment(horizontal="left", vertical="center", indent=1)
    A_LEFT_PLAIN = Alignment(horizontal="left", vertical="center")

    state = {"r": 1}

    def write_main_header(title: str):
        r = state["r"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=title)
        c.font = F_MAIN; c.fill = P_MAIN; c.alignment = A_LEFT
        ws.row_dimensions[r].height = 24
        state["r"] = r + 1

    def write_meta(label: str, value: str):
        r = state["r"]
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = F_LABEL; c1.fill = P_META; c1.alignment = A_LEFT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c2 = ws.cell(row=r, column=2, value=value)
        c2.font = F_VAL; c2.alignment = A_LEFT_PLAIN
        state["r"] = r + 1

    def write_section_bar(title: str):
        # 두 줄 띄우고 네이비 큰 섹션 바 (원본 간격 매치)
        state["r"] += 2
        r = state["r"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=title)
        c.font = F_SECTION; c.fill = P_SECTION; c.alignment = A_LEFT
        ws.row_dimensions[r].height = 22
        state["r"] = r + 2  # 작성 후도 두 줄 빈

    def write_subsection_bar(title: str):
        # 한 줄 띄우고 연파랑 소섹션 바
        state["r"] += 1
        r = state["r"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(row=r, column=1, value=title)
        c.font = F_SUBSEC; c.fill = P_SUBSEC; c.alignment = A_LEFT
        ws.row_dimensions[r].height = 20
        state["r"] = r + 1

    def write_kv(label: str, value):
        # 한 줄 띄우지 않고 바로 (또는 사용 측에서 띄움)
        r = state["r"]
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = F_LABEL; c1.alignment = A_LEFT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c2 = ws.cell(row=r, column=2, value="" if value is None else str(value))
        c2.font = F_VAL; c2.alignment = A_LEFT_PLAIN
        state["r"] = r + 1

    def write_bullet(text: str):
        r = state["r"]
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        c = ws.cell(row=r, column=2, value=f"• {text}")
        c.font = F_BULLET; c.alignment = A_LEFT_PLAIN
        state["r"] = r + 1

    def write_inline_label(label: str):
        """'  CPU 상세 정보:' 같은 작은 인라인 라벨"""
        r = state["r"]
        c = ws.cell(row=r, column=1, value=f"  {label}")
        c.font = F_VAL; c.alignment = A_LEFT_PLAIN
        state["r"] = r + 1

    def gap():
        state["r"] += 1

    # ===== 메인 헤더 + 메타 4행 =====
    write_main_header("iDRAC 조회 결과")
    write_meta("Service Tag", payload.get("__service_tag__", sheet_prefix))
    write_meta("iDRAC IP", ip or "")
    write_meta("조회 옵션", options_str or kind.upper())
    write_meta("조회 시간", timestamp or "")

    # ===== HW =====
    if kind == "hw":
        write_section_bar("Hardware Status && Health Check")

        # System info
        write_subsection_bar("System info")
        gap()
        s = payload.get("system", {})
        write_kv("모델명", s.get("Model"))
        write_kv("제조사", s.get("Manufacturer"))
        write_kv("서비스 태그", s.get("ServiceTag"))
        write_kv("전원 상태", s.get("PowerState"))
        write_kv("시스템 상태", s.get("Health"))

        # Processors
        write_subsection_bar("Processors")
        gap()
        cs = payload.get("cpu_summary", {})
        write_kv("CPU 모델명", cs.get("Model"))
        write_kv("CPU 수량", cs.get("Count"))
        write_kv("CPU 상태", cs.get("Health"))
        write_inline_label("CPU 상세 정보:")
        for c in payload.get("cpus", []):
            write_bullet(f"Socket {c['Socket']}: {c['Speed']} MHz, {c['Cores']} cores")

        # System Memory
        write_subsection_bar("System Memory")
        gap()
        ms = payload.get("memory_summary", {})
        write_kv("Memory 총 용량", f"{ms.get('TotalGiB', 'N/A')} GiB")
        write_kv("Memory 수량", len(payload.get("memory_modules", [])))
        write_kv("Memory 상태", ms.get("Health"))
        write_inline_label("Memory 용량    :")
        for m in payload.get("memory_modules", []):
            write_bullet(f"{m['Location']}: {m['CapacityGB']} GB @ {m['SpeedMHz']} MHz")

        # FAN
        write_subsection_bar("FAN")
        gap()
        fans = payload.get("fans", [])
        write_kv("FAN 수량", len(fans))
        write_kv("FAN 상태",
                 "OK" if fans and all(f.get("Health") == "OK" for f in fans) else
                 ("Not OK" if fans else "N/A"))

        # Storage Controller
        write_subsection_bar("Storage Controller")
        gap()
        for c in payload.get("storage_controllers", []):
            write_bullet(f"{c['Name']} (상태: {c['Health']})")

        # RAID Configuration
        write_subsection_bar("RAID Configuration")
        gap()
        for r2 in payload.get("raids", []):
            write_bullet(f"RAID Level: {r2['RAIDType']}, 용량: {r2['CapacityGB']} GB, 상태: {r2['Health']}")

        # Physical Disks
        write_subsection_bar("Physical Disks")
        gap()
        for d in payload.get("disks", []):
            write_bullet(
                f"Model: {d['Model']}, State: {d['State']}, "
                f"Bus Protocol: {d['Protocol']}, Size: {d['SizeGB']} GB, Status: {d['Health']}"
            )

        # Power Supply
        write_subsection_bar("Power Supply")
        gap()
        ps = payload.get("psu_summary", {})
        write_kv("PSU 수량", ps.get("Count"))
        write_kv("PSU 상태", ps.get("Health"))
        write_kv("PSU 총 전력 용량", f"{ps.get('TotalCapacityW', 'N/A')} W")
        write_kv("PSU 이중화 상태", ps.get("Redundancy"))

        # Network Adapters
        write_subsection_bar("Network Adapters")
        gap()
        for n in payload.get("nics", []):
            write_bullet(
                f"Model: {n['Model']}, Adapter: {n['Adapter']}, "
                f"Port: {n['Port']}, LinkStatus: {n['LinkStatus']}"
            )

    # ===== FW =====
    elif kind == "fw":
        write_section_bar("Firmware Information")

        write_subsection_bar("Service Tag")
        gap()
        write_kv("Service Tag", payload.get("service_tag", "N/A"))

        # show_all 면 all_components 사용, 아니면 핵심 컴포넌트만
        use_all = (payload.get("all_components") and
                   len(payload["all_components"]) > len(payload.get("components", [])))

        if use_all:
            write_subsection_bar(f"Firmware - 전체 컴포넌트 ({len(payload['all_components'])}개)")
            gap()
            for c in payload["all_components"]:
                write_bullet(f"{c['Name']} — {c['Version']}")
        else:
            write_subsection_bar("Firmware (BIOS / LC / iDRAC / PERC / BOSS / CPLD / PSU / TPM 등)")
            gap()
            for c in payload.get("components", []):
                write_bullet(f"{c['Name']} — {c['Version']}")

        write_subsection_bar("NIC Firmware")
        gap()
        for n in payload.get("nic_fw", []):
            write_bullet(f"{n['Slot']} / {n['Model']} — {n['FirmwareVersion']}")

    # ===== BIOS =====
    elif kind == "bios":
        write_section_bar("BIOS Configuration")

        # show_all 모드인 경우 전체 그룹별로
        attrs = payload.get("all_attributes") or {}
        if attrs and len(attrs) > 20:  # show_all 데이터가 풍부하게 있으면
            for name_grp, items in group_bios_attrs(attrs):
                write_subsection_bar(name_grp)
                gap()
                for k, v in items:
                    write_kv(k, v)
            pwr = payload.get("all_power_attributes") or {}
            if pwr:
                write_subsection_bar(f"iDRAC Manager Attributes ({len(pwr)}개)")
                gap()
                for k, v in pwr.items():
                    write_kv(k, v)
        else:
            # 기본 (필터) 모드
            write_subsection_bar("System Setup → System BIOS → System Profile Settings")
            gap()
            for k, v in payload.get("system_profile", {}).items():
                write_kv(k, v)

            write_subsection_bar("System Setup → System BIOS → Processor Settings")
            gap()
            for k, v in payload.get("processor", {}).items():
                write_kv(k, v)

            write_subsection_bar("System Setup → System BIOS → Integrated Devices")
            gap()
            for k, v in payload.get("integrated", {}).items():
                write_kv(k, v)

            write_subsection_bar("System Setup → Power Configuration")
            gap()
            for k, v in payload.get("power_config", {}).items():
                write_kv(k, v)

    wb.save(path)
    return name


# =========================================================
#  Worker Thread
# =========================================================
class FetchWorker(QThread):
    progress = Signal(str)
    finished_ok = Signal(dict)
    failed = Signal(str, str)  # error_kind, message

    def __init__(self, ip, user, pw, modes, sample=False, parent=None):
        super().__init__(parent)
        self.ip, self.user, self.pw = ip, user, pw
        self.modes = modes  # list ['hw'|'fw'|'bios']
        self.sample = sample

    def run(self):
        try:
            if self.sample:
                self.progress.emit("샘플 데이터 생성 중…")
                out = {"service_tag": "SMPL123"}
                if "hw" in self.modes: out["hw"] = sample_payload("hw")
                if "fw" in self.modes: out["fw"] = sample_payload("fw")
                if "bios" in self.modes: out["bios"] = sample_payload("bios")
                self.finished_ok.emit(out)
                return

            insp = Inspector(self.ip, self.user, self.pw)
            out = {}

            self.progress.emit("Service Tag 조회 중…")
            out["service_tag"] = insp.fetch_service_tag()

            if "hw" in self.modes:
                self.progress.emit("HW 정보 조회 중…")
                out["hw"] = insp.fetch_hardware()
                st = out["hw"].get("system", {}).get("ServiceTag")
                if st and st != "N/A":
                    out["service_tag"] = st
            if "fw" in self.modes:
                self.progress.emit("FW 정보 조회 중…")
                out["fw"] = insp.fetch_firmware()
                if out["service_tag"] in ("", "N/A"):
                    st2 = out["fw"].get("service_tag")
                    if st2 and st2 != "N/A":
                        out["service_tag"] = st2
            if "bios" in self.modes:
                self.progress.emit("BIOS 정보 조회 중…")
                out["bios"] = insp.fetch_bios()
            self.finished_ok.emit(out)

        except InspectorError as e:
            if str(e) == "AUTH":
                self.failed.emit("AUTH", "인증 실패 — 사용자/비밀번호를 확인해 주세요.")
            else:
                self.failed.emit("HTTP", f"서버 응답 오류: {e}")
        except requests.exceptions.ConnectTimeout:
            self.failed.emit("TIMEOUT", "연결 시간 초과 — iDRAC IP/네트워크를 확인해 주세요.")
        except requests.exceptions.ConnectionError as e:
            self.failed.emit("NETWORK", f"네트워크 오류 — iDRAC IP({self.ip})에 접근할 수 없습니다.\n{e.__class__.__name__}")
        except requests.exceptions.RequestException as e:
            self.failed.emit("REQ", f"요청 오류: {e}")
        except Exception as e:
            self.failed.emit("UNK", f"예상치 못한 오류:\n{traceback.format_exc()}")


# =========================================================
#  Log Fetch Worker (백그라운드 로그 추출 + 진행률 emit)
# =========================================================
class LogFetchWorker(QThread):
    progress = Signal(int, int, str, str)   # current, total, log_type, log_name
    finished_ok = Signal(dict)              # {'lclog': [...], 'sel': [...]}
    failed = Signal(str, str)
    cancel_requested = False

    def __init__(self, ip, user, pw, log_types, parent=None):
        super().__init__(parent)
        self.ip, self.user, self.pw = ip, user, pw
        self.log_types = log_types     # ['lclog', 'sel']
        self.cancel_requested = False

    def cancel(self):
        self.cancel_requested = True

    def run(self):
        try:
            insp = Inspector(self.ip, self.user, self.pw, timeout=30)
            out = {}
            for lt in self.log_types:
                if self.cancel_requested:
                    break
                name = Inspector.LOG_NAMES.get(lt, lt)
                self.progress.emit(0, 1, lt, name)
                total = insp.fetch_log_count(lt)
                if total == 0:
                    out[lt] = []
                    self.progress.emit(1, 1, lt, name)
                    continue
                entries = []
                skip = 0
                page_size = 50
                while skip < total:
                    if self.cancel_requested:
                        break
                    page = insp.fetch_log_page(lt, skip, page_size)
                    if not page:
                        break
                    entries.extend(page)
                    skip += len(page)
                    self.progress.emit(min(skip, total), total, lt, name)
                out[lt] = entries
            self.finished_ok.emit(out)
        except InspectorError as e:
            if str(e) == "AUTH":
                self.failed.emit("AUTH", "인증 실패 — 사용자/비밀번호를 확인해 주세요.")
            else:
                self.failed.emit("HTTP", f"서버 응답 오류: {e}")
        except requests.exceptions.ConnectTimeout:
            self.failed.emit("TIMEOUT", "연결 시간 초과 — iDRAC IP/네트워크를 확인해 주세요.")
        except requests.exceptions.ConnectionError as e:
            self.failed.emit("NETWORK", f"네트워크 오류 — iDRAC IP({self.ip})에 접근할 수 없습니다.")
        except Exception as e:
            self.failed.emit("UNK", f"예상치 못한 오류:\n{traceback.format_exc()}")


# =========================================================
#  Log Extract / Progress Dialogs
# =========================================================
class LogExtractDialog(QDialog):
    """로그 종류 선택"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("로그 추출")
        self.setMinimumWidth(400)
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        title = QLabel("📥  추출할 iDRAC 로그를 선택하세요")
        title.setStyleSheet("font-size:14px; font-weight:600; color:#1f2328;")
        layout.addWidget(title)

        self.chk_lclog = QCheckBox("Lifecycle Controller Log (LCLog)")
        self.chk_lclog.setChecked(True)
        self.chk_lclog.setToolTip("펌웨어 업데이트, 설정 변경, 부팅, 에러 등 라이프사이클 이벤트")
        layout.addWidget(self.chk_lclog)
        lc_desc = QLabel("    펌웨어 업데이트, 설정 변경, 부팅 / 에러 이력")
        lc_desc.setStyleSheet("color:#57606a; font-size:11px;")
        layout.addWidget(lc_desc)

        self.chk_sel = QCheckBox("System Event Log (SEL)")
        self.chk_sel.setChecked(True)
        self.chk_sel.setToolTip("하드웨어 이벤트 (PSU, FAN, Memory ECC 오류 등)")
        layout.addWidget(self.chk_sel)
        sel_desc = QLabel("    PSU / FAN / Memory ECC 등 하드웨어 이벤트")
        sel_desc.setStyleSheet("color:#57606a; font-size:11px;")
        layout.addWidget(sel_desc)

        layout.addSpacing(10)
        btns = QHBoxLayout()
        btns.addStretch()
        self.cancel_btn = QPushButton("취소")
        self.cancel_btn.clicked.connect(self.reject)
        btns.addWidget(self.cancel_btn)
        self.start_btn = QPushButton("추출 시작")
        self.start_btn.setStyleSheet(
            "QPushButton { background:#0969da; color:white; border:0; "
            "border-radius:6px; padding:8px 20px; font-weight:600; }"
            "QPushButton:hover { background:#0860c9; }")
        self.start_btn.clicked.connect(self.accept)
        btns.addWidget(self.start_btn)
        layout.addLayout(btns)

    def selected_types(self):
        out = []
        if self.chk_lclog.isChecked(): out.append("lclog")
        if self.chk_sel.isChecked(): out.append("sel")
        return out


# =========================================================
#  Firmware Update Worker + Dialogs
# =========================================================
class FirmwareUpdateWorker(QThread):
    """3단계 진행: 1) 파일 업로드  2) iDRAC 검증  3) Job 진행"""
    stage = Signal(str, str)             # stage_key ('upload'|'verify'|'install'), stage_name
    upload_progress = Signal(int, int)   # bytes_sent, total_bytes
    install_progress = Signal(int, str)  # percent, state_message
    finished_ok = Signal(dict)           # {'job_uri': ..., 'apply_time': ..., 'final_state': ...}
    failed = Signal(str, str)
    cancel_requested = False

    def __init__(self, ip, user, pw, file_path, apply_time, parent=None):
        super().__init__(parent)
        self.ip, self.user, self.pw = ip, user, pw
        self.file_path = file_path
        self.apply_time = apply_time   # 'Immediate' or 'OnReset'
        self.cancel_requested = False

    def cancel(self):
        self.cancel_requested = True

    def run(self):
        try:
            insp = Inspector(self.ip, self.user, self.pw, timeout=60)

            # STAGE 1: 서버 상태 확인
            self.stage.emit("verify", "iDRAC 상태 확인 중...")
            ready, reason = insp.check_update_ready()
            if not ready:
                self.failed.emit("BUSY", reason)
                return

            # STAGE 2: 파일 업로드
            self.stage.emit("upload", "펌웨어 파일 업로드 중...")

            def upload_cb(sent, total):
                if not self.cancel_requested:
                    self.upload_progress.emit(sent, total)

            try:
                job_uri = insp.upload_firmware_multipart(
                    self.file_path, self.apply_time, progress_callback=upload_cb
                )
            except Exception as e:
                self.failed.emit("UPLOAD", f"업로드 실패:\n{e}")
                return

            # STAGE 3: Job 진행률 폴링
            self.stage.emit("install", "iDRAC 이 펌웨어 적용 중...")

            import time
            last_state = ""
            stuck_count = 0
            timeout_sec = 3600  # 1시간
            start_ts = time.time()

            while not self.cancel_requested:
                if time.time() - start_ts > timeout_sec:
                    self.failed.emit("TIMEOUT", "1시간이 지나도 Job 이 끝나지 않아 대기 중단.")
                    return

                info = insp.poll_job(job_uri)
                state = info["state"]
                pct = info["percent"]
                msg = info["message"] or f"상태: {state}"

                self.install_progress.emit(pct, msg)

                # 완료 조건
                if state in ("Completed", "RebootCompleted"):
                    self.finished_ok.emit({
                        "job_uri": job_uri,
                        "apply_time": self.apply_time,
                        "final_state": state,
                        "message": msg,
                        "reboot_pending": self.apply_time == "OnReset",
                    })
                    return
                if state in ("Failed", "Exception", "CompletedWithErrors"):
                    self.failed.emit("JOB", f"Job 실패: {state}\n{msg}")
                    return
                # OnReset 모드에서 "Scheduled" 상태면 완료로 간주 (재부팅 대기)
                if self.apply_time == "OnReset" and state in ("Scheduled", "New", "Ready"):
                    if last_state == state:
                        stuck_count += 1
                    else:
                        stuck_count = 0
                    if stuck_count > 5:
                        # 5회 이상 같은 상태면 예약 완료로 간주
                        self.finished_ok.emit({
                            "job_uri": job_uri,
                            "apply_time": self.apply_time,
                            "final_state": state,
                            "message": "다음 재부팅 시 자동 적용됩니다.",
                            "reboot_pending": True,
                        })
                        return

                last_state = state
                time.sleep(5)  # 5초 간격 폴링

            # 취소됨
            self.failed.emit("CANCEL", "사용자가 취소했습니다. iDRAC 쪽 Job 은 여전히 진행 중일 수 있으니 나중에 iDRAC 웹 UI 로 확인해 주세요.")

        except Exception as e:
            self.failed.emit("UNK", f"예상치 못한 오류:\n{traceback.format_exc()}")


class FirmwareSafetyDialog(QDialog):
    """1단계 — 위험 경고 + 3개 안전 확인 체크박스"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⚠️ 펌웨어 업데이트 — 안전 확인")
        self.setMinimumWidth(560)
        layout = QVBoxLayout(self)
        layout.setSpacing(14)

        warn_bg = QLabel(
            "<div style='background:#FFF0F0; border-left:5px solid #D1242F; "
            "padding:12px 16px; color:#9B1C1C; line-height:1.6;'>"
            "<b style='font-size:14px;'>⚠️ 위험한 작업입니다</b><br>"
            "<span style='font-size:11px;'>"
            "• 잘못된 펌웨어를 올리면 서버가 부팅 불가 상태가 될 수 있습니다.<br>"
            "• BIOS/iDRAC 업데이트는 재부팅을 유발할 수 있어 <b>서비스 중단</b> 이 발생합니다.<br>"
            "• 업데이트 중에는 <b>전원을 절대 끄면 안 됩니다.</b>"
            "</span></div>"
        )
        warn_bg.setTextFormat(Qt.RichText)
        warn_bg.setWordWrap(True)
        layout.addWidget(warn_bg)

        confirm_label = QLabel("<b>아래 항목을 모두 확인하셨습니까?</b>")
        confirm_label.setStyleSheet("font-size:13px; color:#1f2328; margin-top:6px;")
        layout.addWidget(confirm_label)

        self.chk1 = QCheckBox("① 지금은 유지보수 시간 (서비스 중이 아님)")
        self.chk2 = QCheckBox("② 대상 서버 데이터 백업이 완료됨 (또는 백업 불필요)")
        self.chk3 = QCheckBox("③ Dell 공식 펌웨어 파일이며 대상 모델과 일치함")
        for c in (self.chk1, self.chk2, self.chk3):
            c.setStyleSheet("font-size:12px; padding:4px 0;")
            layout.addWidget(c)
            c.toggled.connect(self._update_btn)

        layout.addSpacing(6)
        btns = QHBoxLayout()
        btns.addStretch()
        self.cancel_btn = QPushButton("취소")
        self.cancel_btn.clicked.connect(self.reject)
        btns.addWidget(self.cancel_btn)
        self.next_btn = QPushButton("다음 →")
        self.next_btn.setEnabled(False)
        self.next_btn.setStyleSheet(
            "QPushButton { background:#D1242F; color:white; border:0; "
            "border-radius:6px; padding:8px 20px; font-weight:600; }"
            "QPushButton:hover:enabled { background:#B01824; }"
            "QPushButton:disabled { background:#F0C0C0; color:#eef; }")
        self.next_btn.clicked.connect(self.accept)
        btns.addWidget(self.next_btn)
        layout.addLayout(btns)

    def _update_btn(self):
        self.next_btn.setEnabled(self.chk1.isChecked() and self.chk2.isChecked() and self.chk3.isChecked())


class FirmwareSelectDialog(QDialog):
    """2단계 — 파일 선택 + 적용 시점 선택 + 대상 정보"""
    def __init__(self, parent=None, target_info: dict = None):
        super().__init__(parent)
        self.setWindowTitle("펌웨어 파일 선택 및 옵션")
        self.setMinimumWidth(600)
        self.selected_file = None
        self.selected_apply = "Immediate"

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # 대상 서버 정보
        target_info = target_info or {}
        info_html = (
            f"<div style='background:#F6F8FA; border:1px solid #D0D7DE; "
            f"padding:10px 14px; border-radius:6px; font-size:11px;'>"
            f"<b>대상 서버</b><br>"
            f"iDRAC IP: <b>{html.escape(target_info.get('ip', '-'))}</b><br>"
            f"Service Tag: <b>{html.escape(target_info.get('service_tag', '(조회 안 됨)'))}</b>"
            f"</div>"
        )
        lbl_info = QLabel(info_html)
        lbl_info.setTextFormat(Qt.RichText)
        layout.addWidget(lbl_info)

        # 파일 선택
        file_row = QHBoxLayout()
        file_row.addWidget(QLabel("펌웨어 파일:"))
        self.file_input = QLineEdit()
        self.file_input.setReadOnly(True)
        self.file_input.setPlaceholderText("Dell 공식 .exe / .d7 파일 선택...")
        file_row.addWidget(self.file_input, 1)
        self.browse_btn = QPushButton("찾아보기...")
        self.browse_btn.clicked.connect(self._browse_file)
        file_row.addWidget(self.browse_btn)
        layout.addLayout(file_row)

        self.file_info_lbl = QLabel("")
        self.file_info_lbl.setStyleSheet("color:#57606a; font-size:11px; padding-left:4px;")
        layout.addWidget(self.file_info_lbl)

        layout.addSpacing(6)

        # 적용 시점
        apply_group = QLabel("<b>적용 시점</b>")
        apply_group.setStyleSheet("font-size:12px; color:#1f2328;")
        layout.addWidget(apply_group)

        from PySide6.QtWidgets import QRadioButton, QButtonGroup
        self.rb_now = QRadioButton("즉시 적용 (Immediate) — 자동 재부팅 발생, 서비스 중단됨")
        self.rb_now.setStyleSheet("font-size:12px; padding:3px 0;")
        self.rb_now.setChecked(True)
        self.rb_reset = QRadioButton("다음 재부팅 시 적용 (OnReset) — 안전, 유저가 재부팅할 때 반영")
        self.rb_reset.setStyleSheet("font-size:12px; padding:3px 0;")
        layout.addWidget(self.rb_now)
        layout.addWidget(self.rb_reset)

        layout.addSpacing(8)
        btns = QHBoxLayout()
        btns.addStretch()
        self.cancel_btn = QPushButton("취소")
        self.cancel_btn.clicked.connect(self.reject)
        btns.addWidget(self.cancel_btn)
        self.next_btn = QPushButton("다음 →")
        self.next_btn.setEnabled(False)
        self.next_btn.clicked.connect(self._on_next)
        self.next_btn.setStyleSheet(
            "QPushButton { background:#0969DA; color:white; border:0; "
            "border-radius:6px; padding:8px 20px; font-weight:600; }"
            "QPushButton:hover:enabled { background:#0860C9; }"
            "QPushButton:disabled { background:#B0C4DE; color:#eef; }")
        btns.addWidget(self.next_btn)
        layout.addLayout(btns)

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "펌웨어 파일 선택",
            os.path.expanduser("~/Downloads"),
            "Dell 펌웨어 (*.exe *.d7 *.EXE *.D7);;모든 파일 (*.*)"
        )
        if not path:
            return
        self.selected_file = path
        self.file_input.setText(path)
        size_mb = os.path.getsize(path) / (1024 * 1024)
        self.file_info_lbl.setText(f"  파일: {os.path.basename(path)}   크기: {size_mb:.1f} MB")
        self.next_btn.setEnabled(True)

    def _on_next(self):
        self.selected_apply = "Immediate" if self.rb_now.isChecked() else "OnReset"
        self.accept()


class FirmwareConfirmDialog(QDialog):
    """3단계 — 최종 확인 (UPDATE 타이핑 요구)"""
    def __init__(self, parent=None, summary: dict = None):
        super().__init__(parent)
        self.setWindowTitle("최종 확인")
        self.setMinimumWidth(560)
        summary = summary or {}

        layout = QVBoxLayout(self)
        layout.setSpacing(14)

        title = QLabel(
            "<div style='background:#FFF8E1; border-left:5px solid #E8A317; "
            "padding:12px 16px; color:#7A5A00;'>"
            "<b style='font-size:14px;'>🔒 마지막 확인</b><br>"
            "<span style='font-size:11px;'>확인 후에는 되돌릴 수 없습니다.</span>"
            "</div>"
        )
        title.setTextFormat(Qt.RichText)
        layout.addWidget(title)

        # 요약 정보
        sm = (
            f"<table cellpadding='6' style='font-size:12px;'>"
            f"<tr><td style='color:#57606a;'>대상 IP</td><td><b>{html.escape(summary.get('ip', '-'))}</b></td></tr>"
            f"<tr><td style='color:#57606a;'>Service Tag</td><td><b>{html.escape(summary.get('service_tag', '-'))}</b></td></tr>"
            f"<tr><td style='color:#57606a;'>펌웨어 파일</td><td><b>{html.escape(os.path.basename(summary.get('file', '-')))}</b></td></tr>"
            f"<tr><td style='color:#57606a;'>파일 크기</td><td><b>{summary.get('size_mb', 0):.1f} MB</b></td></tr>"
            f"<tr><td style='color:#57606a;'>적용 시점</td><td><b>{summary.get('apply', '-')}</b></td></tr>"
            f"</table>"
        )
        summary_lbl = QLabel(sm)
        summary_lbl.setTextFormat(Qt.RichText)
        layout.addWidget(summary_lbl)

        # UPDATE 타이핑
        confirm_lbl = QLabel(
            "<span style='color:#1f2328; font-size:12px;'>"
            "계속하시려면 아래 칸에 <b><code>UPDATE</code></b> 를 대문자로 정확히 입력하세요:"
            "</span>"
        )
        confirm_lbl.setTextFormat(Qt.RichText)
        layout.addWidget(confirm_lbl)

        self.confirm_input = QLineEdit()
        self.confirm_input.setPlaceholderText("UPDATE")
        self.confirm_input.textChanged.connect(self._check_input)
        layout.addWidget(self.confirm_input)

        layout.addSpacing(6)
        btns = QHBoxLayout()
        btns.addStretch()
        self.cancel_btn = QPushButton("취소")
        self.cancel_btn.clicked.connect(self.reject)
        btns.addWidget(self.cancel_btn)
        self.confirm_btn = QPushButton("업데이트 시작")
        self.confirm_btn.setEnabled(False)
        self.confirm_btn.clicked.connect(self.accept)
        self.confirm_btn.setStyleSheet(
            "QPushButton { background:#D1242F; color:white; border:0; "
            "border-radius:6px; padding:8px 24px; font-weight:700; }"
            "QPushButton:hover:enabled { background:#B01824; }"
            "QPushButton:disabled { background:#F0C0C0; color:#eef; }")
        btns.addWidget(self.confirm_btn)
        layout.addLayout(btns)

    def _check_input(self, text):
        self.confirm_btn.setEnabled(text.strip() == "UPDATE")


class FirmwareProgressDialog(QDialog):
    """진행률 다이얼로그 — 3단계(Verify → Upload → Install)"""
    cancelled = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("펌웨어 업데이트 진행 중")
        self.setMinimumWidth(560)
        self.setModal(True)

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        self.stage_label = QLabel("준비 중...")
        self.stage_label.setStyleSheet("font-size:14px; font-weight:600; color:#1f2328;")
        layout.addWidget(self.stage_label)

        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setTextVisible(True)
        self.progress.setStyleSheet(
            "QProgressBar { border:1px solid #d0d7de; border-radius:4px; "
            "background:#f6f8fa; text-align:center; height:24px; font-weight:600; }"
            "QProgressBar::chunk { background:#0969da; border-radius:3px; }"
        )
        layout.addWidget(self.progress)

        self.detail = QLabel("")
        self.detail.setStyleSheet("color:#57606a; font-size:11px; padding-top:4px;")
        self.detail.setWordWrap(True)
        layout.addWidget(self.detail)

        # 3단계 표시
        self.stages_label = QLabel(
            "<span style='font-size:10px; color:#57606a;'>"
            "① 상태 확인 &nbsp; → &nbsp; ② 파일 업로드 &nbsp; → &nbsp; ③ iDRAC 적용"
            "</span>"
        )
        self.stages_label.setTextFormat(Qt.RichText)
        layout.addWidget(self.stages_label)

        layout.addSpacing(4)
        btns = QHBoxLayout()
        btns.addStretch()
        self.cancel_btn = QPushButton("취소")
        self.cancel_btn.clicked.connect(self._on_cancel)
        btns.addWidget(self.cancel_btn)
        layout.addLayout(btns)

    def _on_cancel(self):
        self.cancelled.emit()
        self.cancel_btn.setText("취소 중...")
        self.cancel_btn.setEnabled(False)

    def set_stage(self, key: str, name: str):
        stage_map = {"verify": "①", "upload": "②", "install": "③"}
        mark = stage_map.get(key, "")
        self.stage_label.setText(f"{mark}  {name}")

    def set_upload_progress(self, sent: int, total: int):
        pct = int(sent * 100 / total) if total > 0 else 0
        self.progress.setValue(pct)
        sent_mb = sent / (1024*1024)
        total_mb = total / (1024*1024)
        self.detail.setText(f"  업로드: {sent_mb:.1f} / {total_mb:.1f} MB ({pct}%)")

    def set_install_progress(self, pct: int, msg: str):
        self.progress.setValue(pct)
        self.detail.setText(f"  {msg}")


class ContactDialog(QDialog):
    """문의 다이얼로그 — 버튼 순서: [GitHub 열기] [URL 복사] [취소]"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("문의하기")
        self.setMinimumWidth(460)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 20, 24, 18)
        layout.setSpacing(14)

        # 아이콘 + 제목
        title = QLabel("💬  문의 / 버그 리포트")
        title.setStyleSheet("font-size:16px; font-weight:700; color:#1f2328;")
        layout.addWidget(title)

        desc = QLabel(
            "버그 리포트, 기능 제안, 개선 요청은 <b>GitHub Issues</b> 로 남겨주세요.<br>"
            "브라우저에서 새 이슈 작성 페이지가 열립니다."
        )
        desc.setTextFormat(Qt.RichText)
        desc.setWordWrap(True)
        desc.setStyleSheet("color:#57606a; font-size:12px; line-height:1.6;")
        layout.addWidget(desc)

        # URL 카드
        url_card = QLabel(f"<div style='color:#0969DA; font-family:Menlo,Consolas,monospace; font-size:11px;'>{ISSUES_URL}</div>")
        url_card.setStyleSheet(
            "background:#F6F8FA; border:1px solid #D0D7DE; border-radius:6px; "
            "padding:10px 12px; color:#0969DA;"
        )
        url_card.setTextFormat(Qt.RichText)
        url_card.setWordWrap(True)
        layout.addWidget(url_card)

        layout.addSpacing(4)

        # 버튼 순서: [GitHub 열기] [URL 복사] [취소]
        btns = QHBoxLayout()
        btns.addStretch()

        self.open_btn = QPushButton("🌐  GitHub 열기")
        self.open_btn.setStyleSheet(
            "QPushButton { background:#0969DA; color:white; border:0; "
            "border-radius:6px; padding:8px 20px; font-weight:600; }"
            "QPushButton:hover { background:#0860C9; }"
            "QPushButton:pressed { background:#0757B9; }"
        )
        self.open_btn.setDefault(True)
        self.open_btn.clicked.connect(self._open_github)
        btns.addWidget(self.open_btn)

        self.copy_btn = QPushButton("📋  URL 복사")
        self.copy_btn.setStyleSheet(
            "QPushButton { background:#F6F8FA; color:#1F2328; border:1px solid #D0D7DE; "
            "border-radius:6px; padding:8px 16px; font-weight:600; }"
            "QPushButton:hover { background:#EAEEF2; }"
        )
        self.copy_btn.clicked.connect(self._copy_url)
        btns.addWidget(self.copy_btn)

        self.cancel_btn = QPushButton("취소")
        self.cancel_btn.setStyleSheet(
            "QPushButton { background:transparent; color:#57606A; border:1px solid #D0D7DE; "
            "border-radius:6px; padding:8px 16px; }"
            "QPushButton:hover { background:#F6F8FA; }"
        )
        self.cancel_btn.clicked.connect(self.reject)
        btns.addWidget(self.cancel_btn)

        layout.addLayout(btns)

    def _open_github(self):
        QDesktopServices.openUrl(QUrl(ISSUES_URL))
        self.accept()

    def _copy_url(self):
        QApplication.clipboard().setText(ISSUES_URL)
        # 임시 라벨로 복사됨 안내
        original = self.copy_btn.text()
        self.copy_btn.setText("✓  복사됨")
        from PySide6.QtCore import QTimer
        QTimer.singleShot(1500, lambda: self.copy_btn.setText(original))


class LogProgressDialog(QDialog):
    """추출 진행률 표시"""
    cancelled = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("로그 추출 중...")
        self.setMinimumWidth(440)
        self.setModal(True)
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        self.label = QLabel("준비 중...")
        self.label.setStyleSheet("font-size:13px; color:#1f2328;")
        layout.addWidget(self.label)

        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setTextVisible(True)
        self.progress.setStyleSheet(
            "QProgressBar { border:1px solid #d0d7de; border-radius:4px; "
            "background:#f6f8fa; text-align:center; height:22px; }"
            "QProgressBar::chunk { background:#0969da; border-radius:3px; }"
        )
        layout.addWidget(self.progress)

        self.detail = QLabel("")
        self.detail.setStyleSheet("color:#57606a; font-size:11px;")
        layout.addWidget(self.detail)

        layout.addSpacing(4)
        btns = QHBoxLayout()
        btns.addStretch()
        self.cancel_btn = QPushButton("취소")
        self.cancel_btn.clicked.connect(self._on_cancel)
        btns.addWidget(self.cancel_btn)
        layout.addLayout(btns)

    def _on_cancel(self):
        self.cancelled.emit()
        self.cancel_btn.setText("취소 중...")
        self.cancel_btn.setEnabled(False)

    def update_progress(self, current: int, total: int, log_type: str, log_name: str):
        pct = int(current * 100 / total) if total > 0 else 0
        self.label.setText(f"📥 {log_name}")
        self.detail.setText(f"  {current:,} / {total:,} 항목 ({pct}%)")
        self.progress.setValue(pct)


# =========================================================
#  Main Window
# =========================================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"{APP_NAME}  v{APP_VERSION}")
        self.resize(1180, 820)
        self.settings = QSettings("longchiri", "iDRACViewer")
        self.last_payload: dict = {}
        self.worker: Optional[FetchWorker] = None

        # 창 아이콘
        icon_path = resource_path("iDRAC_Viewer.icns") or resource_path("iDRAC_Viewer.png") or resource_path("iDRAC_Viewer.ico")
        if icon_path:
            self.setWindowIcon(QIcon(icon_path))

        self._build_ui()
        self._restore_settings()

    # ---------- UI ----------
    def _build_ui(self):
        # 최상위: 탭 위젯
        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)
        self.tabs.setStyleSheet(
            "QTabWidget::pane { border: 1px solid #d0d7de; border-radius:8px; background:white; top:-1px; }"
            "QTabBar::tab { background:#f6f8fa; color:#57606a; padding:10px 24px; "
            "  border:1px solid #d0d7de; border-bottom:0; "
            "  border-top-left-radius:8px; border-top-right-radius:8px; "
            "  margin-right:2px; font-size:13px; font-weight:600; }"
            "QTabBar::tab:selected { background:white; color:#0969DA; }"
            "QTabBar::tab:hover:!selected { background:#eaeef2; }"
        )
        self.setCentralWidget(self.tabs)

        # ── 탭 1: 서버 정보 ──
        server_tab = QWidget()
        root = QVBoxLayout(server_tab)
        root.setContentsMargins(14, 14, 14, 10)
        root.setSpacing(10)

        # 접속 정보
        conn_box = QGroupBox("iDRAC 접속 정보")
        g = QGridLayout(conn_box)
        g.setHorizontalSpacing(10); g.setVerticalSpacing(8)

        g.addWidget(QLabel("iDRAC IP"), 0, 0)
        self.ip_in = QLineEdit(); self.ip_in.setPlaceholderText("예: 192.168.0.120")
        g.addWidget(self.ip_in, 0, 1)

        g.addWidget(QLabel("Username"), 0, 2)
        self.user_in = QLineEdit("root"); self.user_in.setMaximumWidth(160)
        g.addWidget(self.user_in, 0, 3)

        g.addWidget(QLabel("Password"), 0, 4)
        pw_wrap = QHBoxLayout()
        self.pw_in = QLineEdit(); self.pw_in.setEchoMode(QLineEdit.Password)
        self.pw_in.setMinimumWidth(220)
        pw_wrap.addWidget(self.pw_in)
        self.pw_show = QCheckBox("보기")
        self.pw_show.toggled.connect(
            lambda on: self.pw_in.setEchoMode(QLineEdit.Normal if on else QLineEdit.Password))
        pw_wrap.addWidget(self.pw_show)
        pw_widget = QWidget(); pw_widget.setLayout(pw_wrap)
        g.addWidget(pw_widget, 0, 5)

        g.addWidget(QLabel("조회 항목"), 1, 0)
        mode_widget = QWidget()
        mode_layout = QHBoxLayout(mode_widget)
        mode_layout.setContentsMargins(0, 0, 0, 0)
        mode_layout.setSpacing(12)

        self.chk_hw = QCheckBox("HW")
        self.chk_hw.setChecked(True)
        self.chk_fw = QCheckBox("FW")
        self.chk_bios = QCheckBox("BIOS")
        mode_layout.addWidget(self.chk_hw)
        mode_layout.addWidget(self.chk_fw)
        mode_layout.addWidget(self.chk_bios)

        sep1 = QFrame(); sep1.setFrameShape(QFrame.VLine); sep1.setStyleSheet("color:#ccc;")
        mode_layout.addWidget(sep1)

        self.chk_show_all = QCheckBox("모든 항목 표시")
        self.chk_show_all.setToolTip("BIOS/FW를 필터링하지 않고 전체 값을 그룹별로 보여줍니다")
        mode_layout.addWidget(self.chk_show_all)

        self.chk_5g = QCheckBox("5g 모드")
        self.chk_5g.setToolTip("운영에서 자주 확인하는 핵심 BIOS 항목만 추려서 표시 (모든 항목 표시 OFF일 때만 적용)")
        mode_layout.addWidget(self.chk_5g)

        mode_layout.addStretch()
        g.addWidget(mode_widget, 1, 1, 1, 2)

        # 상호 배제: show_all 켜지면 5g 비활성
        self.chk_show_all.toggled.connect(lambda on: (self.chk_5g.setDisabled(on),
                                                     self.chk_5g.setChecked(False) if on else None))

        self.run_btn = QPushButton("▶  실행")
        self.run_btn.setMinimumHeight(36)
        self.run_btn.setMinimumWidth(110)
        self.run_btn.setStyleSheet(
            "QPushButton { background:#0969da; color:white; border:0; border-radius:6px; padding:6px 22px; font-weight:600; font-size:13px; }"
            "QPushButton:hover { background:#0860c9; }"
            "QPushButton:pressed { background:#0757b9; }"
            "QPushButton:disabled { background:#a3b8d4; color:#eef; }")
        self.run_btn.clicked.connect(self.on_run)
        g.addWidget(self.run_btn, 1, 4)

        self.sample_btn = QPushButton("샘플 보기")
        self.sample_btn.setMinimumHeight(36)
        self.sample_btn.setMinimumWidth(110)
        self.sample_btn.clicked.connect(self.on_sample)
        g.addWidget(self.sample_btn, 1, 5)

        root.addWidget(conn_box)

        # 결과 위 도구 바 (두 줄)
        # ── 1줄: 표시 옵션 (보기좋게 / 폰트 / 키필터) ──
        tools_row1 = QHBoxLayout()
        tools_row1.setSpacing(10)

        self.pretty_chk = QCheckBox("보기 좋게")
        self.pretty_chk.setChecked(True)
        self.pretty_chk.toggled.connect(self._rerender)
        tools_row1.addWidget(self.pretty_chk)

        tools = tools_row1  # 기존 코드에서 tools 참조하는 부분 유지

        # show_all / 5g 토글이 바뀌면 같은 데이터로 즉시 재렌더
        self.chk_show_all.toggled.connect(self._rerender)
        self.chk_5g.toggled.connect(self._rerender)

        tools.addSpacing(12)
        tools.addWidget(QLabel("폰트"))
        self.font_spin = QSpinBox(); self.font_spin.setRange(8, 24); self.font_spin.setValue(12)
        self.font_spin.valueChanged.connect(self._apply_font_size)
        tools.addWidget(self.font_spin)

        tools.addSpacing(16)
        tools.addWidget(QLabel("🔍 키 필터"))
        self.key_filter_in = QLineEdit()
        self.key_filter_in.setPlaceholderText("예: Proc*, MemFreq, Boot   (콤마/공백 구분, * 와일드카드)")
        self.key_filter_in.setMinimumWidth(280)
        self.key_filter_in.setToolTip(
            "BIOS 속성/FW 컴포넌트를 키워드로 필터링합니다.\n"
            "  • 콤마/공백으로 여러 키워드 입력\n"
            "  • 대소문자 무시, 부분 일치\n"
            "  • '*' 와일드카드 지원 (예: Proc*, Mem* )\n"
            "  • 비워두면 필터 OFF\n"
            "  • show_all 켜고 사용하면 가장 잘 동작함"
        )
        self.key_filter_in.textChanged.connect(self._rerender)
        tools.addWidget(self.key_filter_in)

        self.key_filter_clear_btn = QPushButton("✕")
        self.key_filter_clear_btn.setMaximumWidth(28)
        self.key_filter_clear_btn.setToolTip("필터 지우기")
        self.key_filter_clear_btn.clicked.connect(lambda: self.key_filter_in.clear())
        tools.addWidget(self.key_filter_clear_btn)

        tools.addStretch()
        root.addLayout(tools_row1)

        # ── 2줄: 액션 버튼들 (텍스트 액션 | 저장 | 위험 액션 | 유틸리티) ──
        tools_row2 = QHBoxLayout()
        tools_row2.setSpacing(6)

        self.copy_all_btn = QPushButton("전체 선택")
        self.copy_all_btn.clicked.connect(lambda: (self.result.selectAll(), self.result.copy()))
        tools_row2.addWidget(self.copy_all_btn)

        self.copy_btn = QPushButton("복사")
        self.copy_btn.clicked.connect(lambda: self.result.copy())
        tools_row2.addWidget(self.copy_btn)

        self.save_txt_btn = QPushButton("저장 (TXT)")
        self.save_txt_btn.clicked.connect(self.on_save_txt)
        tools_row2.addWidget(self.save_txt_btn)

        # 그룹 구분자 1
        sep_a = QFrame(); sep_a.setFrameShape(QFrame.VLine); sep_a.setStyleSheet("color:#D0D7DE;")
        sep_a.setFixedHeight(24)
        tools_row2.addSpacing(6)
        tools_row2.addWidget(sep_a)
        tools_row2.addSpacing(6)
        # tools 변수를 row2로 이관
        tools = tools_row2

        self.save_xlsx_btn = QPushButton("📊  엑셀로 저장")
        self.save_xlsx_btn.setStyleSheet(
            "QPushButton { background:#1a7f37; color:white; border:0; border-radius:6px; padding:6px 14px; font-weight:600; }"
            "QPushButton:hover { background:#15722f; }"
            "QPushButton:pressed { background:#116628; }"
            "QPushButton:disabled { background:#a8d4b6; color:#eef; }")
        self.save_xlsx_btn.clicked.connect(self.on_save_xlsx)
        tools.addWidget(self.save_xlsx_btn)

        self.log_btn = QPushButton("📥  로그 추출")
        self.log_btn.setStyleSheet(
            "QPushButton { background:#6f42c1; color:white; border:0; border-radius:6px; padding:6px 14px; font-weight:600; }"
            "QPushButton:hover { background:#5a2da3; }"
            "QPushButton:pressed { background:#4c2588; }"
            "QPushButton:disabled { background:#c4b3e0; color:#eef; }")
        self.log_btn.setToolTip("iDRAC LCLog / SEL 로그 추출 후 엑셀로 저장")
        self.log_btn.clicked.connect(self.on_log_extract)
        tools.addWidget(self.log_btn)

        self.fw_btn = QPushButton("🔧  펌웨어 업데이트")
        self.fw_btn.setStyleSheet(
            "QPushButton { background:#D1242F; color:white; border:0; border-radius:6px; padding:6px 14px; font-weight:600; }"
            "QPushButton:hover { background:#B01824; }"
            "QPushButton:pressed { background:#8E0F1B; }"
            "QPushButton:disabled { background:#F0C0C0; color:#eef; }")
        self.fw_btn.setToolTip("⚠️ 위험한 작업 — 3단계 안전 확인 후 진행")
        self.fw_btn.clicked.connect(self.on_firmware_update)
        tools.addWidget(self.fw_btn)

        # 그룹 구분자 2 (액션 vs 유틸리티)
        tools.addStretch()
        sep_b = QFrame(); sep_b.setFrameShape(QFrame.VLine); sep_b.setStyleSheet("color:#D0D7DE;")
        sep_b.setFixedHeight(24)
        tools.addWidget(sep_b)
        tools.addSpacing(6)

        # 문의 버튼 (도구바 오른쪽 끝)
        self.contact_btn = QPushButton("❓ 문의")
        self.contact_btn.setStyleSheet(
            "QPushButton { background:#f6f8fa; color:#0969da; border:1px solid #d0d7de; "
            "border-radius:6px; padding:6px 12px; font-weight:600; }"
            "QPushButton:hover { background:#eaeef2; }"
            "QPushButton:pressed { background:#d0d7de; }")
        self.contact_btn.setToolTip("버그 리포트 / 개선 요청 (GitHub Issues)")
        self.contact_btn.clicked.connect(self.on_contact)
        tools.addWidget(self.contact_btn)

        self.clear_btn = QPushButton("지우기")
        self.clear_btn.clicked.connect(self._clear)
        tools.addWidget(self.clear_btn)

        root.addLayout(tools_row2)

        # 결과창
        self.result = QTextBrowser()
        self.result.setOpenExternalLinks(False)
        self._apply_font_size(self.font_spin.value())
        root.addWidget(self.result, 1)

        # 상태바
        self.status = QStatusBar()
        self.setStatusBar(self.status)
        self.status.showMessage("대기 중")

        # 서버 탭 등록
        self.tabs.addTab(server_tab, "🖥  서버 정보")

        # ── 탭 2: 사용법 ──
        help_tab = self._build_help_tab()
        self.tabs.addTab(help_tab, "📖  사용법")

    def _build_help_tab(self):
        wrap = QWidget()
        layout = QVBoxLayout(wrap)
        layout.setContentsMargins(0, 0, 0, 0)

        help_view = QTextBrowser()
        help_view.setOpenExternalLinks(True)
        help_view.setStyleSheet(
            "QTextBrowser { background:#F6F8FA; border:0; padding:0; }"
        )
        help_view.document().setDefaultStyleSheet(
            "body { font-family: -apple-system, BlinkMacSystemFont, 'Malgun Gothic', sans-serif; }"
        )
        help_view.setHtml(self._help_html())
        layout.addWidget(help_view)
        return wrap

    def _help_html_OLD(self):
        """(구버전 — 사용 안 함)"""
        return ""

    def _help_html(self):
        """앱 내장 사용법 — Qt QTextBrowser 호환 table 기반 카드 UI
           (max-width 강제, 카드형 섹션, 아이콘 뱃지)
        """
        # ─── 스타일 정의 ───
        # 색깔
        C_BG = "#F6F8FA"
        C_CARD = "#FFFFFF"
        C_TEXT = "#1F2328"
        C_TEXT2 = "#57606A"
        C_BLUE = "#0969DA"
        C_NAVY = "#264F78"
        C_LIGHTBLUE = "#DDF4FF"
        C_BORDER = "#D0D7DE"
        C_ORANGE = "#E8A317"
        C_DANGER = "#D1242F"
        C_DANGER_BG = "#FFEBE9"
        C_TIP_BG = "#DDF4FF"
        C_TIP_FG = "#0550AE"
        C_WARN_BG = "#FFF8E1"
        C_WARN_FG = "#7A5A00"

        # 헬퍼 — 큰 섹션 (번호 원 + 제목)
        def section(num, title):
            return (
                f'<table cellpadding="0" cellspacing="0" style="margin:40px 0 12px 0;">'
                f'<tr>'
                f'<td valign="middle" style="background:{C_BLUE}; color:white; '
                f'width:36px; height:36px; text-align:center; '
                f'font-size:16px; font-weight:800; border-radius:18px;">{num}</td>'
                f'<td valign="middle" style="padding-left:14px; color:{C_TEXT}; '
                f'font-size:20px; font-weight:700; letter-spacing:-0.3px;">{title}</td>'
                f'</tr>'
                f'</table>'
            )

        # 헬퍼 — 소섹션 (연파랑 좌측바 + 파란 텍스트)
        def subsection(title):
            return (
                f'<table cellpadding="0" cellspacing="0" width="100%" style="margin:22px 0 10px 0;">'
                f'<tr>'
                f'<td width="4" bgcolor="{C_BLUE}" style="background:{C_BLUE};"> </td>'
                f'<td style="padding:2px 0 2px 12px; color:{C_BLUE}; '
                f'font-size:14px; font-weight:700;">{title}</td>'
                f'</tr>'
                f'</table>'
            )

        # 헬퍼 — 카드 (배경 흰색 + 얇은 테두리)
        def card(inner_html):
            return (
                f'<table cellpadding="0" cellspacing="0" width="100%" '
                f'style="margin:8px 0 12px 0;">'
                f'<tr>'
                f'<td bgcolor="{C_CARD}" style="background:{C_CARD}; '
                f'border:1px solid {C_BORDER}; border-radius:8px; padding:16px 20px;">'
                f'{inner_html}</td>'
                f'</tr></table>'
            )

        # 헬퍼 — 팁/경고/위험 박스
        def box(bg_color, fg_color, border_color, icon, title, body):
            return (
                f'<table cellpadding="0" cellspacing="0" width="100%" '
                f'style="margin:14px 0;">'
                f'<tr>'
                f'<td width="4" bgcolor="{border_color}" style="background:{border_color};"> </td>'
                f'<td bgcolor="{bg_color}" style="background:{bg_color}; '
                f'padding:14px 18px; color:{fg_color};">'
                f'<div style="font-weight:700; font-size:13px; margin-bottom:6px;">{icon}&nbsp; {title}</div>'
                f'<div style="font-size:12.5px; line-height:1.7;">{body}</div>'
                f'</td></tr></table>'
            )

        # 헬퍼 — 정의 리스트 (라벨 : 값)
        def def_list(pairs):
            rows = ""
            for label, val in pairs:
                rows += (
                    f'<tr>'
                    f'<td valign="top" style="padding:5px 16px 5px 0; '
                    f'color:{C_TEXT}; font-weight:700; font-size:13px; white-space:nowrap;">{label}</td>'
                    f'<td valign="top" style="padding:5px 0; color:{C_TEXT2}; '
                    f'font-size:13px; line-height:1.7;">{val}</td>'
                    f'</tr>'
                )
            return f'<table cellpadding="0" cellspacing="0">{rows}</table>'

        # 헬퍼 — 아이콘 리스트 (아이콘 + 라벨 + 설명)
        def icon_list(items):
            # items = [(emoji, label, desc), ...]
            rows = ""
            for i, (emoji, label, desc) in enumerate(items):
                border = f'border-top:1px solid #EAEDF0;' if i > 0 else ''
                rows += (
                    f'<tr>'
                    f'<td valign="top" style="width:36px; padding:12px 6px 12px 0; '
                    f'font-size:18px; {border}">{emoji}</td>'
                    f'<td valign="top" style="padding:12px 12px 12px 0; '
                    f'color:{C_TEXT}; font-weight:700; font-size:13px; width:90px; {border}">{label}</td>'
                    f'<td valign="top" style="padding:12px 0; color:{C_TEXT2}; '
                    f'font-size:12.5px; line-height:1.6; {border}">{desc}</td>'
                    f'</tr>'
                )
            return f'<table cellpadding="0" cellspacing="0" width="100%">{rows}</table>'

        # 헬퍼 — 코드 인라인
        def code(text):
            return (f'<span style="background:#EFF1F3; color:{C_BLUE}; '
                    f'padding:2px 7px; border-radius:3px; font-family:Menlo,Consolas,monospace; '
                    f'font-size:12px;">{text}</span>')

        # ============================================================
        # 콘텐츠 조립
        # ============================================================

        # ── 타이틀 헤더 (그라디언트 느낌 카드) ──
        header = (
            f'<table cellpadding="0" cellspacing="0" width="100%">'
            f'<tr>'
            f'<td bgcolor="{C_NAVY}" style="background:{C_NAVY}; '
            f'padding:26px 28px; border-radius:8px;">'
            f'<div style="color:white; font-size:26px; font-weight:800; '
            f'letter-spacing:-0.5px;">📖  iDRAC Toolkit 사용법</div>'
            f'<div style="color:#B6D4F0; font-size:12.5px; margin-top:6px;">'
            f'v{APP_VERSION} &nbsp;·&nbsp; Dell PowerEdge 서버 관리 도구</div>'
            f'</td></tr></table>'
        )

        # ── 3초 요약 ──
        tldr = box(
            C_TIP_BG, C_TIP_FG, C_BLUE, "🚀", "3초 요약",
            f"IP·계정 입력 &nbsp;→&nbsp; 조회 항목 체크 &nbsp;→&nbsp; "
            f"<b>[▶ 실행]</b> &nbsp;→&nbsp; 결과 확인 &nbsp;→&nbsp; 필요시 저장"
        )

        # ── 1. 앱 소개 ──
        s1_content = icon_list([
            ("🖥", "HW", "모델, CPU, 메모리, 팬, 스토리지, RAID, 디스크, PSU, NIC"),
            ("🔧", "FW", "BIOS, iDRAC, LC, PERC, BOSS, CPLD, PSU, TPM 등"),
            ("⚙️", "BIOS 설정", "System Profile, Processor, Integrated Devices, Power"),
            ("📥", "로그 추출", "LCLog / SEL 을 컬러 엑셀로 저장"),
            ("🔧", "펌웨어 업데이트", "3단계 안전 확인 후 진행"),
        ])
        s1 = section("1", "앱 소개") + card(s1_content)

        # ── 2. 기본 사용법 ──
        s2_1 = subsection("① 접속 정보 입력") + card(def_list([
            ("iDRAC IP", "서버 iDRAC 관리 IP"),
            ("Username", f"기본값 {code('root')}"),
            ("Password", f"{code('☑ 보기')} 체크로 확인 가능"),
        ]))
        s2_2 = subsection("② 조회 항목 선택") + card(def_list([
            ("HW / FW / BIOS", "중복 선택 가능 (예: HW+FW 만)"),
            ("모든 항목 표시", "BIOS 전체 속성 (150~475개) 그룹별 표시"),
            ("5g 모드", "운영 핵심 BIOS 15개 항목만 프리셋"),
        ]))
        s2_3 = subsection("③ [▶ 실행] 클릭") + card(
            f'<div style="color:{C_TEXT2}; font-size:13px; line-height:1.8;">'
            f'백그라운드에서 조회 → 결과창에 컬러로 정리되어 표시됩니다.<br>'
            f'<b>Service Tag / IP / 옵션 / 조회 시간</b>은 항상 상단에 표시됩니다.'
            f'</div>'
        )
        s2 = section("2", "기본 사용법") + s2_1 + s2_2 + s2_3

        # ── 3. 키 필터 ──
        rules = card(def_list([
            ("대소문자 무시", f"{code('proc')} = {code('Proc')} = {code('PROC')}"),
            ("부분 일치", f"{code('Mem')} → Memory, MemFrequency, MemPatrolScrub"),
            ("여러 개", f"콤마/공백 구분 — {code('Mem, Boot, Pxe')}"),
            ("와일드카드 *", f"{code('Proc*')} = \"Proc로 시작\" 만"),
        ]))

        # 키 필터 예시 표 (더 예쁘게)
        examples_data = [
            ("Mem*, Dimm*, Numa", "메모리 관련 모두"),
            ("Proc*, Cpu*, Turbo", "CPU/프로세서 관련"),
            ("Boot*, Uefi, SecureBoot", "부팅 관련"),
            ("Pxe*, Network*, Http*", "네트워크 / PXE"),
            ("Tpm*, Sec*, Aes", "보안 / TPM"),
            ("Power*, AcPwr*, Energy", "전원"),
            ("Virt*, Sriov, Iommu", "가상화"),
            ("BIOS, iDRAC, PERC, BOSS", "특정 펌웨어만"),
            ("NIC, Broadcom, Intel", "NIC 펌웨어"),
            ("Disk*, Drive", "디스크 펌웨어"),
        ]
        ex_rows = ""
        for i, (k, v) in enumerate(examples_data):
            bg = "#F6F8FA" if i % 2 == 1 else "#FFFFFF"
            ex_rows += (
                f'<tr>'
                f'<td bgcolor="{bg}" style="background:{bg}; padding:8px 14px; '
                f'font-family:Menlo,Consolas,monospace; font-size:12px; '
                f'color:{C_BLUE}; width:50%;">{k}</td>'
                f'<td bgcolor="{bg}" style="background:{bg}; padding:8px 14px; '
                f'color:{C_TEXT2}; font-size:12.5px;">{v}</td>'
                f'</tr>'
            )
        ex_table = (
            f'<table cellpadding="0" cellspacing="0" width="100%" '
            f'style="border:1px solid {C_BORDER}; border-radius:6px;">'
            f'<tr><td bgcolor="{C_NAVY}" style="background:{C_NAVY}; color:white; '
            f'padding:8px 14px; font-size:12px; font-weight:700; width:50%;">입력 키워드</td>'
            f'<td bgcolor="{C_NAVY}" style="background:{C_NAVY}; color:white; '
            f'padding:8px 14px; font-size:12px; font-weight:700;">추출 항목</td></tr>'
            f'{ex_rows}'
            f'</table>'
        )

        tip_filter = box(
            C_TIP_BG, C_TIP_FG, C_BLUE, "💡", "가장 잘 쓰는 방법",
            f"<b>모든 항목 표시</b> 체크박스를 같이 켜면 검색 범위가 BIOS 전체로 넓어져 "
            f"필터가 가장 잘 동작합니다."
        )

        s3 = (
            section("3", "🔍 키 필터 사용법") +
            f'<div style="color:{C_TEXT2}; font-size:13px; line-height:1.8; margin-bottom:10px;">'
            f'결과창 위 <b>키 필터</b> 입력칸에 키워드를 넣으면 매칭되는 BIOS 속성 / FW 컴포넌트만 표시됩니다.'
            f'</div>' +
            subsection("규칙 4가지") + rules +
            subsection("바로 복사해 쓰는 예시 10가지") +
            f'<div style="margin:8px 0 14px 0;">{ex_table}</div>' +
            tip_filter
        )

        # ── 4. 엑셀 저장 ──
        s4 = section("4", "📊 엑셀로 저장") + card(
            f'<div style="color:{C_TEXT}; font-size:13px; line-height:1.8;">'
            f'결과창 위 <b>[📊 엑셀로 저장]</b> 클릭 → 저장 위치 선택 → 완료.'
            f'</div>' +
            def_list([
                ("시트명", f"<b>Service Tag 기준</b> 예: {code('ABCD123')}"),
                ("통합", "HW / FW / BIOS 모두 <b>한 시트에 통합</b> 저장"),
                ("누적", "같은 파일을 다시 선택하면 <b>다른 서버 시트가 추가</b>"),
                ("디자인", "앱 화면과 동일 톤 (진파랑 / 네이비 / 연파랑)"),
            ])
        )

        # ── 5. 로그 추출 ──
        s5 = section("5", "📥 로그 추출 (LCLog / SEL)") + card(
            f'<div style="color:{C_TEXT}; font-size:13px; line-height:1.8;">'
            f'<b>[📥 로그 추출]</b> → 로그 종류 선택 → 진행률 표시 → 저장 위치 지정.'
            f'</div>' +
            def_list([
                ("LCLog", "펌웨어 업데이트, 설정 변경, 부팅, 에러 이력"),
                ("SEL", "PSU / FAN / Memory ECC 등 하드웨어 이벤트"),
                ("색상", "🔴 Critical &nbsp; 🟡 Warning &nbsp; 🟢 OK"),
                ("취소", "진행 중에도 취소 가능"),
            ])
        )

        # ── 6. 펌웨어 업데이트 ──
        danger = box(
            C_DANGER_BG, C_DANGER, C_DANGER, "⚠️", "위험한 작업입니다",
            "• 잘못된 펌웨어 → 부팅 불가 상태 가능<br>"
            "• BIOS 업데이트는 자동 재부팅 → 서비스 중단<br>"
            "• 업데이트 중 <b>전원 절대 금지</b>"
        )
        fw_steps = card(
            f'<div style="color:{C_TEXT}; font-size:13px; font-weight:700; margin-bottom:12px;">'
            f'4단계 안전 프로세스</div>' +
            icon_list([
                ("①", "위험 경고", "3개 체크박스 확인 — 유지보수 시간 / 백업 완료 / Dell 공식 펌웨어"),
                ("②", "파일 선택", f"적용 시점 선택 — 즉시({code('Immediate')}) 또는 다음 재부팅 시({code('OnReset')})"),
                ("③", "최종 확인", f"{code('UPDATE')} 대문자 타이핑 요구"),
                ("④", "진행률", "파일 업로드 → iDRAC 적용 (5초 간격 폴링)"),
            ])
        )
        s6 = section("6", "🔧 펌웨어 업데이트") + danger + fw_steps

        # ── 7. 지원 모델 ──
        models_data = [
            ("14G", "R640, R740, R740xd, R840, R940"),
            ("15G", "R650, R750, R6515, R7515, R6525, R7525"),
            ("16G", "R660, R760, R6615, R7615, R6625, R7625"),
            ("블레이드", "MX740c, MX750c, MX760c"),
        ]
        m_rows = ""
        for i, (gen, mods) in enumerate(models_data):
            bg = "#F6F8FA" if i % 2 == 1 else "#FFFFFF"
            m_rows += (
                f'<tr>'
                f'<td bgcolor="{bg}" style="background:{bg}; padding:10px 16px; '
                f'color:{C_BLUE}; font-weight:700; font-size:12.5px; width:80px;">{gen}</td>'
                f'<td bgcolor="{bg}" style="background:{bg}; padding:10px 16px; '
                f'color:{C_TEXT}; font-size:12.5px;">{mods}</td>'
                f'</tr>'
            )
        m_table = (
            f'<table cellpadding="0" cellspacing="0" width="100%" '
            f'style="border:1px solid {C_BORDER}; border-radius:6px; margin-top:8px;">'
            f'<tr><td bgcolor="{C_NAVY}" style="background:{C_NAVY}; color:white; '
            f'padding:8px 16px; font-size:12px; font-weight:700;">세대</td>'
            f'<td bgcolor="{C_NAVY}" style="background:{C_NAVY}; color:white; '
            f'padding:8px 16px; font-size:12px; font-weight:700;">모델 예시</td></tr>'
            f'{m_rows}</table>'
        )
        s7 = section("7", "지원 모델") + (
            f'<div style="color:{C_TEXT2}; font-size:13px; margin-bottom:8px;">'
            f'표준 Redfish API를 쓰는 <b>모든 Dell PowerEdge 모델 자동 감지</b>'
            f'</div>{m_table}'
        )

        # ── 8. FAQ ──
        faqs = [
            ("한글 비밀번호로 로그인이 안 돼요",
             "v3.0.0부터 UTF-8 Basic Auth 로 변경되어 한글 / 특수문자 비밀번호도 지원합니다."),
            ("HTTP 503 응답을 받았어요",
             "iDRAC 이 일시적 응답 불가 상태입니다. 부팅 중이거나 펌웨어 업데이트 중일 수 있어요. 1~2분 기다린 후 다시 [실행]."),
            ("결과창이 비어 있어요",
             "체크박스 (HW / FW / BIOS) 중 최소 하나는 켜야 합니다."),
            ("키 필터를 입력했는데 0개 나와요",
             "기본 모드에서는 14개 핵심 항목 안에서만 찾아서 매칭이 적습니다. [모든 항목 표시]를 켜고 다시 시도하세요."),
            ("macOS \"손상됨\" 경고가 떠요",
             f"터미널: {code('xattr -cr &quot;/Applications/iDRAC Toolkit.app&quot;')}<br>"
             f"또는 앱과 함께 배포된 <b>install_mac.command</b> 스크립트를 더블클릭"),
            ("Windows SmartScreen 경고가 떠요",
             "\"추가 정보\" → \"실행\" 클릭 한 번만 하면 됩니다. "
             "zip 안의 <b>install_win.bat</b> 을 더블클릭하면 자동으로 Program Files 이동 + "
             "바탕화면 바로가기 생성까지 됩니다."),
            ("Windows Defender 가 .exe 를 삭제/격리했어요",
             "PyInstaller 앱은 종종 오탐지됩니다. Windows 보안 → 바이러스 및 위협 방지 → "
             "보호 기록 에서 항목 <b>[복원]</b> 클릭. "
             "또는 install_win.bat 스크립트 안에 자세한 예외 등록 방법이 표시됩니다."),
        ]
        faq_cards = ""
        for q, a in faqs:
            faq_cards += card(
                f'<div style="color:{C_TEXT}; font-weight:700; font-size:13.5px; '
                f'margin-bottom:6px;">Q. {q}</div>'
                f'<div style="color:{C_TEXT2}; font-size:12.5px; line-height:1.7;">A. {a}</div>'
            )
        s8 = section("8", "문제 해결 (FAQ)") + faq_cards

        # ── 푸터 ──
        footer = (
            f'<div style="text-align:center; color:{C_TEXT2}; font-size:11px; '
            f'padding:44px 0 12px 0; margin-top:32px; border-top:1px solid #EAEDF0;">'
            f'📮 &nbsp;버그 리포트 · 기능 제안은 도구바 <b>[❓ 문의]</b> 버튼<br><br>'
            f'<span style="font-size:10px; color:#8C959F;">'
            f'iDRAC Toolkit v{APP_VERSION} &nbsp;·&nbsp; Made by longchiri</span>'
            f'</div>'
        )

        # 모든 콘텐츠 하나로
        body_content = header + tldr + s1 + s2 + s3 + s4 + s5 + s6 + s7 + s8 + footer

        # ── 최외곽 table: 최대폭 720 + 중앙 정렬 + 좌우 여백 ──
        return f"""
<table cellpadding="0" cellspacing="0" width="100%" bgcolor="{C_BG}">
  <tr>
    <td align="center" style="padding:24px 12px 60px 12px;">
      <table cellpadding="0" cellspacing="0" width="720" style="width:720px;">
        <tr>
          <td>
            {body_content}
          </td>
        </tr>
      </table>
    </td>
  </tr>
</table>
"""
        # 스타일 상수 — 톤 다운
        H1 = ("color:#1F2328; font-size:22px; font-weight:700; "
              "margin:44px 0 8px 0; padding-bottom:8px; "
              "border-bottom:2px solid #E8A317;")
        H2 = ("color:#0969DA; font-size:15px; font-weight:700; "
              "margin:26px 0 8px 0; padding-left:12px; "
              "border-left:3px solid #0969DA;")
        H3 = ("color:#1F2328; font-size:13px; font-weight:700; "
              "margin:16px 0 6px 0;")
        P  = "color:#333; font-size:13px; line-height:1.8; margin:6px 0;"
        LI = "color:#333; font-size:13px; line-height:1.9;"

        BOX_TIP = ("background:#DDF4FF; border:1px solid #B6E3FF; "
                   "border-left:4px solid #0969DA; "
                   "padding:14px 18px; margin:14px 0; "
                   "color:#0550AE; font-size:12.5px; line-height:1.7; border-radius:4px;")
        BOX_WARN = ("background:#FFF8E1; border:1px solid #F0C36D; "
                    "border-left:4px solid #E8A317; "
                    "padding:14px 18px; margin:14px 0; "
                    "color:#7A5A00; font-size:12.5px; line-height:1.7; border-radius:4px;")
        BOX_DANGER = ("background:#FFEBE9; border:1px solid #FF8182; "
                      "border-left:4px solid #D1242F; "
                      "padding:14px 18px; margin:14px 0; "
                      "color:#9B1C1C; font-size:12.5px; line-height:1.7; border-radius:4px;")

        CODE = ("background:#EFF1F3; color:#0969DA; "
                "padding:2px 6px; border-radius:3px; "
                "font-family:Menlo,Consolas,monospace; font-size:12px;")

        TABLE = ("border-collapse:collapse; width:100%; margin:14px 0; "
                 "font-size:12.5px; border-radius:4px; overflow:hidden;")
        TH = ("background:#0E639C; color:white; text-align:left; "
              "padding:10px 14px; font-weight:600; font-size:12px;")
        TD = ("padding:9px 14px; border-bottom:1px solid #EAEDF0; color:#333;")

        # 최상위 컨테이너 — 최대폭 720px, 여백 넉넉
        wrapper_open = ("<div style='max-width:720px; margin:0 auto; "
                        "padding:20px 8px 60px 8px; font-family:-apple-system,BlinkMacSystemFont,"
                        "\"Helvetica Neue\",sans-serif;'>")
        wrapper_close = "</div>"

        return f"""
{wrapper_open}
<div style="padding-bottom:24px; margin-bottom:8px; border-bottom:1px solid #EAEDF0;">
  <div style="font-size:28px; font-weight:800; color:#1F2328; letter-spacing:-0.5px;">
    📖 iDRAC Toolkit 사용법
  </div>
  <div style="color:#57606A; font-size:12px; margin-top:4px;">
    v{APP_VERSION} &nbsp;·&nbsp; Dell PowerEdge 서버 관리 도구
  </div>
</div>

<div style="{BOX_TIP}">
  🚀 <b>3초 요약</b><br>
  IP·계정 입력 → 조회 항목 체크 → <b>[▶ 실행]</b> → 결과 확인 → 필요시 저장
</div>

<div style="{H1}">1. 앱 소개</div>
<p style="{P}">
  Dell PowerEdge 서버의 <b>iDRAC Redfish API</b> 를 통해 HW / FW / BIOS 정보를 조회하고,
  로그 추출·펌웨어 업데이트까지 한 화면에서 처리하는 도구입니다.
</p>
<ul style="{LI}">
  <li><b>HW</b> — 모델, CPU, 메모리, 팬, 스토리지, RAID, 디스크, PSU, NIC</li>
  <li><b>FW</b> — BIOS, iDRAC, LC, PERC, BOSS, CPLD, PSU, TPM 등</li>
  <li><b>BIOS 설정</b> — System Profile, Processor, Integrated Devices, Power</li>
  <li><b>로그 추출</b> — LCLog / SEL 을 엑셀로 저장</li>
  <li><b>펌웨어 업데이트</b> — 3단계 안전 확인 후 진행</li>
</ul>

<div style="{H1}">2. 기본 사용법</div>

<div style="{H2}">① 접속 정보 입력</div>
<ul style="{LI}">
  <li><b>iDRAC IP</b> &nbsp;— 서버 iDRAC 관리 IP</li>
  <li><b>Username</b> &nbsp;— 기본값 <span style="{CODE}">root</span></li>
  <li><b>Password</b> &nbsp;— <b>[보기]</b> 체크로 확인 가능</li>
</ul>

<div style="{H2}">② 조회 항목 선택</div>
<ul style="{LI}">
  <li><b>HW / FW / BIOS</b> &nbsp;— 중복 선택 가능 (예: HW+FW 만)</li>
  <li><b>모든 항목 표시</b> &nbsp;— BIOS 전체 속성(150~475개) 그룹별 표시</li>
  <li><b>5g 모드</b> &nbsp;— 운영 핵심 BIOS 15개만 프리셋</li>
</ul>

<div style="{H2}">③ [▶ 실행] 클릭</div>
<p style="{P}">
  백그라운드에서 조회 → 결과창에 컬러로 정리되어 표시됩니다.<br>
  Service Tag / IP / 옵션 / 조회 시간은 <b>항상 상단</b>에 표시됩니다.
</p>

<div style="{H1}">3. 🔍 키 필터 사용법</div>
<p style="{P}">
  결과창 위 <b>키 필터</b> 입력칸에 키워드를 넣으면 매칭되는 BIOS 속성 / FW 컴포넌트만 표시됩니다.
</p>

<div style="{H2}">규칙</div>
<ul style="{LI}">
  <li><b>대소문자 무시</b> &nbsp;<span style="{CODE}">proc</span> = <span style="{CODE}">Proc</span> = <span style="{CODE}">PROC</span></li>
  <li><b>부분 일치</b> &nbsp;<span style="{CODE}">Mem</span> → Memory, MemFrequency, MemPatrolScrub</li>
  <li><b>여러 개</b> &nbsp;콤마/공백 구분 — <span style="{CODE}">Mem, Boot, Pxe</span></li>
  <li><b>와일드카드 *</b> &nbsp;<span style="{CODE}">Proc*</span> = "Proc로 시작"만</li>
</ul>

<div style="{H2}">바로 복사해 쓰는 예시</div>
<table style="{TABLE}">
  <tr><th style="{TH}">입력 키워드</th><th style="{TH}">추출 항목</th></tr>
  <tr><td style="{TD} font-family:Menlo,monospace; color:#0969DA;">Mem*, Dimm*, Numa</td><td style="{TD}">메모리 관련 모두</td></tr>
  <tr><td style="{TD} font-family:Menlo,monospace; color:#0969DA; background:#F6F8FA;">Proc*, Cpu*, Turbo</td><td style="{TD} background:#F6F8FA;">CPU/프로세서 관련</td></tr>
  <tr><td style="{TD} font-family:Menlo,monospace; color:#0969DA;">Boot*, Uefi, SecureBoot</td><td style="{TD}">부팅 관련</td></tr>
  <tr><td style="{TD} font-family:Menlo,monospace; color:#0969DA; background:#F6F8FA;">Pxe*, Network*, Http*</td><td style="{TD} background:#F6F8FA;">네트워크 / PXE</td></tr>
  <tr><td style="{TD} font-family:Menlo,monospace; color:#0969DA;">Tpm*, Sec*, Aes</td><td style="{TD}">보안 / TPM</td></tr>
  <tr><td style="{TD} font-family:Menlo,monospace; color:#0969DA; background:#F6F8FA;">Power*, AcPwr*, Energy</td><td style="{TD} background:#F6F8FA;">전원</td></tr>
  <tr><td style="{TD} font-family:Menlo,monospace; color:#0969DA;">Virt*, Sriov, Iommu</td><td style="{TD}">가상화</td></tr>
  <tr><td style="{TD} font-family:Menlo,monospace; color:#0969DA; background:#F6F8FA;">BIOS, iDRAC, PERC, BOSS</td><td style="{TD} background:#F6F8FA;">특정 펌웨어만</td></tr>
  <tr><td style="{TD} font-family:Menlo,monospace; color:#0969DA;">NIC, Broadcom, Intel</td><td style="{TD}">NIC 펌웨어</td></tr>
  <tr><td style="{TD} font-family:Menlo,monospace; color:#0969DA; background:#F6F8FA;">Disk*, Drive</td><td style="{TD} background:#F6F8FA;">디스크 펌웨어</td></tr>
</table>

<div style="{BOX_TIP}">
  💡 <b>모든 항목 표시</b> 체크박스를 같이 켜면 검색 범위가 BIOS 전체로 넓어져 필터가 가장 잘 동작합니다.
</div>

<div style="{H1}">4. 📊 엑셀로 저장</div>
<p style="{P}">
  결과창 위 <b>[📊 엑셀로 저장]</b> 클릭 → 저장 위치 선택 → 완료.
</p>
<ul style="{LI}">
  <li>시트명은 <b>Service Tag 기준</b> — 예: <span style="{CODE}">ABCD123</span></li>
  <li>HW / FW / BIOS 모두 <b>한 시트에 통합</b> 저장</li>
  <li>같은 파일을 다시 선택하면 다른 서버 결과가 <b>추가 시트로 누적</b></li>
  <li>디자인 톤은 앱 화면과 동일 (진파랑 / 네이비 / 연파랑)</li>
</ul>

<div style="{H1}">5. 📥 로그 추출 (LCLog / SEL)</div>
<p style="{P}">
  <b>[📥 로그 추출]</b> 버튼 → 로그 종류 선택 → 진행률 표시 → 저장 위치 지정.
</p>
<ul style="{LI}">
  <li><b>LCLog</b> &nbsp;— 펌웨어 업데이트, 설정 변경, 부팅, 에러 이력</li>
  <li><b>SEL</b> &nbsp;— PSU / FAN / Memory ECC 등 하드웨어 이벤트</li>
  <li>Severity 별 <b>자동 색상</b> — 🔴 Critical / 🟡 Warning / 🟢 OK</li>
  <li>진행률 실시간 표시, 취소 가능</li>
</ul>

<div style="{H1}">6. 🔧 펌웨어 업데이트</div>
<div style="{BOX_DANGER}">
  ⚠️ <b>위험한 작업입니다.</b><br>
  • 잘못된 펌웨어 → 부팅 불가 상태 가능<br>
  • BIOS 업데이트는 자동 재부팅 → 서비스 중단<br>
  • 업데이트 중 <b>전원 절대 금지</b>
</div>
<p style="{P}"><b>4단계 안전 프로세스</b>:</p>
<ol style="{LI}">
  <li><b>위험 경고 + 3개 체크</b> — 유지보수 시간 / 백업 완료 / Dell 공식 펌웨어 확인</li>
  <li><b>파일 선택 + 적용 시점</b> — 즉시(Immediate) 또는 다음 재부팅 시(OnReset)</li>
  <li><b>최종 확인</b> — <span style="{CODE}">UPDATE</span> 대문자 타이핑</li>
  <li><b>진행률 표시</b> — 파일 업로드 → iDRAC 적용 (5초 간격 폴링)</li>
</ol>

<div style="{H1}">7. 지원 모델</div>
<p style="{P}">
  표준 Redfish API 를 쓰는 <b>모든 Dell PowerEdge 모델 자동 감지</b>
</p>
<table style="{TABLE}">
  <tr><th style="{TH}" width="90">세대</th><th style="{TH}">모델 예시</th></tr>
  <tr><td style="{TD} color:#0969DA; font-weight:600;">14G</td><td style="{TD}">R640, R740, R740xd, R840, R940</td></tr>
  <tr><td style="{TD} color:#0969DA; font-weight:600; background:#F6F8FA;">15G</td><td style="{TD} background:#F6F8FA;">R650, R750, R6515, R7515, R6525, R7525</td></tr>
  <tr><td style="{TD} color:#0969DA; font-weight:600;">16G</td><td style="{TD}">R660, R760, R6615, R7615, R6625, R7625</td></tr>
  <tr><td style="{TD} color:#0969DA; font-weight:600; background:#F6F8FA;">블레이드</td><td style="{TD} background:#F6F8FA;">MX740c, MX750c, MX760c</td></tr>
</table>

<div style="{H1}">8. 문제 해결 (FAQ)</div>

<div style="{H3}">Q. 한글 비밀번호로 로그인이 안 돼요</div>
<p style="{P}">
  v3.0.0부터 UTF-8 Basic Auth 로 변경되어 한글 / 특수문자 비밀번호도 지원합니다.
</p>

<div style="{H3}">Q. HTTP 503 응답을 받았어요</div>
<p style="{P}">
  iDRAC 이 일시적 응답 불가 상태입니다. 부팅 중이거나 펌웨어 업데이트 중일 수 있어요.<br>
  1~2분 기다린 후 다시 <b>[실행]</b>.
</p>

<div style="{H3}">Q. 결과창이 비어 있어요</div>
<p style="{P}">체크박스 (HW / FW / BIOS) 중 최소 하나는 켜야 합니다.</p>

<div style="{H3}">Q. 키 필터를 입력했는데 0개 나와요</div>
<p style="{P}">
  기본 모드에서는 14개 핵심 항목 안에서만 찾아서 매칭이 적습니다.<br>
  <b>[모든 항목 표시]</b> 를 켜고 다시 시도하세요.
</p>

<div style="{H3}">Q. macOS "손상됨" 경고가 떠요</div>
<p style="{P}">터미널: <span style="{CODE}">xattr -cr "/Applications/iDRAC Toolkit.app"</span></p>

<div style="{H3}">Q. Windows SmartScreen 경고가 떠요</div>
<p style="{P}">"추가 정보" → "실행" 클릭 한 번만 하면 됩니다.</p>

<div style="text-align:center; color:#57606A; font-size:11px; padding:40px 0 10px 0; margin-top:40px; border-top:1px solid #EAEDF0;">
  📮 버그 리포트 / 기능 제안은 도구바 <b>[❓ 문의]</b> 버튼<br><br>
  <span style="font-size:10px; color:#8C959F;">iDRAC Toolkit v{APP_VERSION} · Made by longchiri</span>
</div>
{wrapper_close}
"""

    # ---------- 설정 저장/복원 ----------
    def _restore_settings(self):
        self.ip_in.setText(self.settings.value("ip", "", str))
        u = self.settings.value("user", "root", str)
        self.user_in.setText(u)
        self.font_spin.setValue(int(self.settings.value("font", 12)))
        # 체크박스 상태 복원 (기본: HW만 ON)
        self.chk_hw.setChecked(self.settings.value("chk_hw", True, type=bool))
        self.chk_fw.setChecked(self.settings.value("chk_fw", False, type=bool))
        self.chk_bios.setChecked(self.settings.value("chk_bios", False, type=bool))
        self.chk_show_all.setChecked(self.settings.value("show_all", False, type=bool))
        self.chk_5g.setChecked(self.settings.value("preset_5g", False, type=bool))
        if self.chk_show_all.isChecked():
            self.chk_5g.setDisabled(True)
        # 키 필터 복원
        self.key_filter_in.setText(self.settings.value("key_filter", "", str))
        last_xlsx = self.settings.value("last_xlsx", "", str)
        self._last_xlsx = last_xlsx if last_xlsx else ""
        # 초기 웰컴 화면
        self._show_welcome()

    def closeEvent(self, event):
        self.settings.setValue("ip", self.ip_in.text())
        self.settings.setValue("user", self.user_in.text())
        self.settings.setValue("font", self.font_spin.value())
        self.settings.setValue("chk_hw", self.chk_hw.isChecked())
        self.settings.setValue("chk_fw", self.chk_fw.isChecked())
        self.settings.setValue("chk_bios", self.chk_bios.isChecked())
        self.settings.setValue("show_all", self.chk_show_all.isChecked())
        self.settings.setValue("preset_5g", self.chk_5g.isChecked())
        self.settings.setValue("key_filter", self.key_filter_in.text())
        self.settings.setValue("last_xlsx", getattr(self, "_last_xlsx", ""))
        super().closeEvent(event)

    # ---------- 실행 ----------
    def on_run(self): self._run(sample=False)
    def on_sample(self): self._run(sample=True)

    def _run(self, sample: bool):
        if self.worker and self.worker.isRunning():
            QMessageBox.information(self, APP_NAME, "이전 작업이 끝난 뒤 다시 시도해 주세요.")
            return

        ip = self.ip_in.text().strip()
        user = self.user_in.text().strip()
        pw = self.pw_in.text()

        # 체크박스 → modes 리스트
        modes = []
        if self.chk_hw.isChecked(): modes.append("hw")
        if self.chk_fw.isChecked(): modes.append("fw")
        if self.chk_bios.isChecked(): modes.append("bios")

        if not modes:
            QMessageBox.warning(self, APP_NAME, "조회할 항목(HW/FW/BIOS) 중 최소 하나를 선택해 주세요.")
            return

        if not sample:
            missing = []
            if not ip: missing.append("iDRAC IP")
            if not user: missing.append("Username")
            if not pw: missing.append("Password")
            if missing:
                QMessageBox.warning(self, APP_NAME, "다음 항목을 입력해 주세요:\n• " + "\n• ".join(missing))
                return

        self._set_busy(True)
        self.result.clear()
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        mode_str = " + ".join([m.upper() for m in modes])
        extra = []
        if self.chk_show_all.isChecked(): extra.append("모든 항목")
        if self.chk_5g.isChecked(): extra.append("5g 모드")
        if extra: mode_str += f" ({', '.join(extra)})"
        head = f"[{ts}] {'(샘플) ' if sample else ''}대상: {ip or '(없음)'}  /  옵션: {mode_str}"
        self._append_html(f'<div style="color:#666; font-size:11px;">{html.escape(head)}</div>')

        self.worker = FetchWorker(ip, user, pw, modes, sample=sample)
        self.worker.progress.connect(self.status.showMessage)
        self.worker.finished_ok.connect(self._on_done)
        self.worker.failed.connect(self._on_fail)
        self.worker.start()

    def _on_done(self, payload: dict):
        self.last_payload = payload
        self._render(payload)
        self._set_busy(False)
        self.status.showMessage("완료 ✓", 5000)

    def _on_fail(self, kind: str, msg: str):
        self._set_busy(False)
        self.status.showMessage(f"실패 ({kind})", 8000)
        title = {"AUTH": "인증 실패",
                 "TIMEOUT": "연결 시간 초과",
                 "NETWORK": "네트워크 오류",
                 "HTTP": "서버 응답 오류"}.get(kind, "오류")
        QMessageBox.critical(self, title, msg)

    def _set_busy(self, busy):
        for w in (self.run_btn, self.sample_btn, self.ip_in, self.user_in, self.pw_in,
                  self.chk_hw, self.chk_fw, self.chk_bios, self.chk_show_all, self.chk_5g):
            w.setEnabled(not busy)
        # show_all 우선 — busy 풀린 뒤에도 상호 배제 유지
        if not busy and self.chk_show_all.isChecked():
            self.chk_5g.setDisabled(True)
        self.run_btn.setText("조회 중…" if busy else "실행")

    def _current_mode_text(self) -> str:
        """현재 체크된 모드를 사람이 읽을 수 있는 문자열로"""
        modes = []
        if self.chk_hw.isChecked(): modes.append("HW")
        if self.chk_fw.isChecked(): modes.append("FW")
        if self.chk_bios.isChecked(): modes.append("BIOS")
        s = " + ".join(modes) if modes else "-"
        extra = []
        if self.chk_show_all.isChecked(): extra.append("모든 항목")
        if self.chk_5g.isChecked() and not self.chk_show_all.isChecked(): extra.append("5g 모드")
        if extra: s += f" ({', '.join(extra)})"
        return s

    # ---------- 렌더 ----------
    def _render(self, payload: dict):
        st = payload.get("service_tag", "N/A") or "N/A"
        ip = self.ip_in.text().strip() or "(샘플)"
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if self.pretty_chk.isChecked():
            # ---- 엑셀과 동일한 메인 헤더 + 메타 4행 ----
            html_out = (
                # 메인 헤더 바 — 엑셀의 #0E639C
                '<div style="background:#0E639C; color:white; '
                'padding:12px 18px; margin:0 0 0 0; '
                'font-weight:700; font-size:17px; letter-spacing:0.5px;">'
                'iDRAC 조회 결과</div>'
                # 메타 정보 4행 — 엑셀의 #EEEEEE 라벨 / 값 회색 텍스트
                '<table cellspacing="0" cellpadding="0" '
                'style="border-collapse:collapse; width:100%; margin:0 0 18px 0; '
                'border:1px solid #ddd; border-top:0;">'
                f'<tr><td style="background:#EEEEEE; padding:7px 12px; '
                f'font-weight:700; color:#333; width:140px; border-bottom:1px solid #e5e5e5;">Service Tag</td>'
                f'<td style="padding:7px 12px; color:#333; font-weight:700; '
                f'letter-spacing:2px; font-family:Menlo,Consolas,monospace; '
                f'border-bottom:1px solid #e5e5e5;">{html.escape(str(st))}</td></tr>'
                f'<tr><td style="background:#EEEEEE; padding:7px 12px; '
                f'font-weight:700; color:#333; border-bottom:1px solid #e5e5e5;">iDRAC IP</td>'
                f'<td style="padding:7px 12px; color:#333; '
                f'border-bottom:1px solid #e5e5e5;">{html.escape(ip)}</td></tr>'
                f'<tr><td style="background:#EEEEEE; padding:7px 12px; '
                f'font-weight:700; color:#333; border-bottom:1px solid #e5e5e5;">조회 옵션</td>'
                f'<td style="padding:7px 12px; color:#333; '
                f'border-bottom:1px solid #e5e5e5;">{html.escape(self._current_mode_text())}</td></tr>'
                f'<tr><td style="background:#EEEEEE; padding:7px 12px; '
                f'font-weight:700; color:#333;">조회 시간</td>'
                f'<td style="padding:7px 12px; color:#333;">{html.escape(ts)}</td></tr>'
                '</table>'
            )

            show_all = self.chk_show_all.isChecked()
            preset_5g = self.chk_5g.isChecked() and not show_all
            key_filter = parse_key_filter(self.key_filter_in.text())
            if "hw" in payload:
                html_out += html_main_section("Hardware Status && Health Check")
                html_out += format_hw_html(payload["hw"])
            if "fw" in payload:
                html_out += html_main_section("Firmware Information")
                html_out += format_fw_html(payload["fw"], show_all=show_all, key_filter=key_filter)
            if "bios" in payload:
                html_out += html_main_section("BIOS Configuration")
                html_out += format_bios_html(payload["bios"], show_all=show_all, preset_5g=preset_5g, key_filter=key_filter)
            self.result.setHtml(html_out)
        else:
            show_all = self.chk_show_all.isChecked()
            preset_5g = self.chk_5g.isChecked() and not show_all
            key_filter = parse_key_filter(self.key_filter_in.text())
            sep = "=" * 60
            txt_out = [sep,
                       f"  Service Tag : {st}",
                       f"  대상 IP     : {ip}",
                       f"  조회 일시   : {ts}",
                       sep]
            if "hw" in payload:
                txt_out.append("\n#### HW ####"); txt_out.append(format_hw_text(payload["hw"]))
            if "fw" in payload:
                txt_out.append("\n#### FW ####"); txt_out.append(format_fw_text(payload["fw"], show_all=show_all, key_filter=key_filter))
            if "bios" in payload:
                txt_out.append("\n#### BIOS ####"); txt_out.append(format_bios_text(payload["bios"], show_all=show_all, preset_5g=preset_5g, key_filter=key_filter))
            self.result.setPlainText("\n".join(txt_out))

    def _rerender(self):
        if self.last_payload:
            self._render(self.last_payload)

    def _append_html(self, html_str: str):
        self.result.append(html_str)

    def _clear(self):
        self.last_payload = {}
        self._show_welcome()
        self.status.showMessage("결과 지움")

    def _show_welcome(self):
        """빈 결과창 안내"""
        self.result.setHtml(f"""
<div style="text-align:center; padding:60px 20px; color:#57606a;">
  <div style="font-size:48px; margin-bottom:14px;">🖥️</div>
  <div style="font-size:18px; color:#1f2328; font-weight:700; margin-bottom:6px;">
    iDRAC Toolkit v{APP_VERSION}
  </div>
  <div style="font-size:12px; margin-bottom:24px;">Dell PowerEdge 서버 관리 도구</div>

  <div style="max-width:520px; margin:0 auto; text-align:left; padding:18px 22px;
              background:#f6f8fa; border:1px solid #d0d7de; border-radius:8px; font-size:12px;">
    <div style="color:#0969DA; font-weight:700; margin-bottom:10px;">👉 시작하기</div>
    <ol style="margin:0; padding-left:20px; line-height:1.9;">
      <li>위쪽 <b>iDRAC 접속 정보</b> 입력 (IP / Username / Password)</li>
      <li><b>조회 항목</b> 체크 (HW / FW / BIOS)</li>
      <li>파란 <b>[▶ 실행]</b> 버튼 클릭</li>
      <li>결과가 이 화면에 표시됩니다</li>
    </ol>
    <div style="margin-top:14px; padding-top:12px; border-top:1px solid #d0d7de; color:#57606a;">
      💡 실제 서버 없이 미리 보고 싶으면 <b>[샘플 보기]</b> 버튼<br>
      📖 자세한 설명은 상단 <b>[사용법]</b> 탭
    </div>
  </div>
</div>
""")


    def _apply_font_size(self, sz: int):
        f = self.result.font()
        f.setPointSize(sz)
        self.result.setFont(f)

    # ---------- 저장 ----------
    def on_save_txt(self):
        if not self.last_payload:
            QMessageBox.information(self, APP_NAME, "저장할 결과가 없습니다.")
            return
        default = f"dell_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        path, _ = QFileDialog.getSaveFileName(self, "TXT 저장", default, "Text (*.txt);;All (*.*)")
        if not path: return
        try:
            with open(path, "w", encoding="utf-8") as f:
                if self.pretty_chk.isChecked():
                    f.write(self.result.toPlainText())
                else:
                    f.write(self.result.toPlainText())
            self.status.showMessage(f"저장됨: {path}", 5000)
        except Exception as e:
            QMessageBox.critical(self, "저장 실패", str(e))

    def on_save_xlsx(self):
        if not self.last_payload:
            QMessageBox.information(self, APP_NAME, "저장할 결과가 없습니다. 먼저 [실행]으로 조회해 주세요.")
            return

        # Service Tag 추출 (시트 이름 접두어)
        tag = self.last_payload.get("service_tag", "N/A")
        if (not tag) or tag == "N/A":
            # 후보 fallback
            if "hw" in self.last_payload:
                tag = self.last_payload["hw"].get("system", {}).get("ServiceTag", "N/A")
            elif "fw" in self.last_payload:
                tag = self.last_payload["fw"].get("service_tag", "N/A")
        tag = (tag or "N/A").replace(" ", "_").replace("/", "_")[:20]

        initial = self._last_xlsx or os.path.join(
            os.path.expanduser("~"), "Documents",
            f"dell_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        path, _ = QFileDialog.getSaveFileName(
            self, "Excel 저장 (기존 파일 선택 시 시트가 추가됩니다)",
            initial, "Excel (*.xlsx)"
        )
        if not path: return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        # 메타 정보 수집 (시트 상단에 IP/옵션/시간 표시)
        ip = self.ip_in.text().strip() or "(샘플)"
        modes_for_meta = []
        if "hw" in self.last_payload: modes_for_meta.append("HW")
        if "fw" in self.last_payload: modes_for_meta.append("FW")
        if "bios" in self.last_payload: modes_for_meta.append("BIOS")
        options_str = " + ".join(modes_for_meta)
        if self.chk_show_all.isChecked(): options_str += " (모든 항목)"
        elif self.chk_5g.isChecked(): options_str += " (5g 모드)"
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # HW/FW/BIOS 를 모두 하나의 시트에 통합 저장 (시트명 = Service Tag)
        payloads = {}
        if "hw" in self.last_payload: payloads["hw"] = self.last_payload["hw"]
        if "fw" in self.last_payload: payloads["fw"] = self.last_payload["fw"]
        if "bios" in self.last_payload: payloads["bios"] = self.last_payload["bios"]

        used = []
        try:
            used.append(save_combined_xlsx(path, tag, payloads,
                                           ip=ip, options_str=options_str, timestamp=ts))
        except Exception as e:
            QMessageBox.critical(self, "엑셀 저장 실패", f"{e}\n\n{traceback.format_exc()}")
            return

        self._last_xlsx = path
        QMessageBox.information(
            self, "저장 완료",
            f"파일: {path}\n추가된 시트: {', '.join(used)}\n\n"
            "같은 파일을 다시 선택해서 다른 서버 결과를 누적 저장할 수 있습니다."
        )

    # ============================================================
    # 📥 로그 추출
    # ============================================================
    def on_log_extract(self):
        # 입력 검증
        ip = self.ip_in.text().strip()
        user = self.user_in.text().strip()
        pw = self.pw_in.text()
        missing = []
        if not ip: missing.append("iDRAC IP")
        if not user: missing.append("Username")
        if not pw: missing.append("Password")
        if missing:
            QMessageBox.warning(self, APP_NAME,
                "로그 추출에는 iDRAC 접속 정보가 필요합니다:\n• " + "\n• ".join(missing))
            return

        # 1) 로그 종류 선택 다이얼로그
        sel_dlg = LogExtractDialog(self)
        if sel_dlg.exec() != QDialog.Accepted:
            return
        log_types = sel_dlg.selected_types()
        if not log_types:
            QMessageBox.information(self, APP_NAME, "로그 종류를 최소 하나 선택해 주세요.")
            return

        # 2) 진행률 다이얼로그
        self.log_progress_dlg = LogProgressDialog(self)

        # 3) Worker 시작
        self.log_worker = LogFetchWorker(ip, user, pw, log_types)
        self.log_worker.progress.connect(self.log_progress_dlg.update_progress)
        self.log_worker.finished_ok.connect(self._on_log_done)
        self.log_worker.failed.connect(self._on_log_fail)
        self.log_progress_dlg.cancelled.connect(self.log_worker.cancel)

        self.log_btn.setEnabled(False)
        self.log_worker.start()
        self.log_progress_dlg.exec()  # 모달 표시

    def _on_log_done(self, payload: dict):
        self.log_progress_dlg.close()
        self.log_btn.setEnabled(True)

        # 결과 요약
        counts = {k: len(v) for k, v in payload.items()}
        if not any(counts.values()):
            QMessageBox.information(self, APP_NAME, "추출된 로그가 없습니다.")
            return

        summary = "\n".join([f"  • {Inspector.LOG_NAMES.get(k, k)}: {c:,}개"
                             for k, c in counts.items()])

        # 4) 저장 위치 선택 다이얼로그
        st = (self.last_payload or {}).get("service_tag", "")
        if not st or st == "N/A":
            try:
                # 빠르게 ServiceTag 만 다시 조회 (없을 경우)
                from datetime import datetime as _dt
                st = "Server"
            except Exception:
                st = "Server"
        default_name = f"iDRAC_Logs_{st}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        default_dir = self.settings.value("last_log_dir",
                                          os.path.join(os.path.expanduser("~"), "Documents"), str)
        initial = os.path.join(default_dir, default_name)

        path, _ = QFileDialog.getSaveFileName(
            self,
            f"로그 저장 위치 선택  ({summary.strip()})",
            initial,
            "Excel 파일 (*.xlsx);;CSV 파일 (*.csv)"
        )
        if not path:
            return  # 사용자가 취소

        if not path.lower().endswith((".xlsx", ".csv")):
            path += ".xlsx"

        # 저장
        try:
            if path.lower().endswith(".csv"):
                self._save_logs_csv(path, payload, st)
            else:
                self._save_logs_xlsx(path, payload, st)
            self.settings.setValue("last_log_dir", os.path.dirname(path))
            QMessageBox.information(self, "저장 완료",
                f"파일: {path}\n\n포함된 로그:\n{summary}")
        except Exception as e:
            QMessageBox.critical(self, "저장 실패", f"{e}\n\n{traceback.format_exc()}")

    def _on_log_fail(self, kind: str, msg: str):
        if hasattr(self, 'log_progress_dlg'):
            self.log_progress_dlg.close()
        self.log_btn.setEnabled(True)
        title = {"AUTH": "인증 실패", "TIMEOUT": "연결 시간 초과",
                 "NETWORK": "네트워크 오류", "HTTP": "서버 응답 오류"}.get(kind, "오류")
        QMessageBox.critical(self, title, msg)

    def _save_logs_xlsx(self, path: str, payload: dict, service_tag: str):
        """추출된 로그를 색상 있는 엑셀로 저장 (로그 종류별 시트)"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        F_HEAD = Font(bold=True, color=XLSX_COLOR_WHITE, size=11)
        P_HEAD = PatternFill("solid", fgColor=XLSX_COLOR_SECTION_BG)
        # Severity 별 색상
        SEV_FILL = {
            "Critical": PatternFill("solid", fgColor="FFCCCC"),
            "Warning":  PatternFill("solid", fgColor="FFF3CD"),
            "OK":       PatternFill("solid", fgColor="D4F4DD"),
        }
        SEV_FONT = {
            "Critical": Font(color="9B1C1C", bold=True),
            "Warning":  Font(color="8C6D1F"),
            "OK":       Font(color="1A7F37"),
        }

        for lt, entries in payload.items():
            sheet_name = ("LCLog" if lt == "lclog" else "SEL")[:31]
            ws = wb.create_sheet(sheet_name)
            ws.column_dimensions["A"].width = 8
            ws.column_dimensions["B"].width = 11
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["E"].width = 70

            # 헤더
            headers = ["#", "Severity", "Created", "MessageId", "Message"]
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.font = F_HEAD; c.fill = P_HEAD
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 22

            # 데이터
            for i, e in enumerate(entries, 1):
                sev = e.get("Severity", "")
                ws.cell(row=i+1, column=1, value=i)
                ws.cell(row=i+1, column=2, value=sev)
                ws.cell(row=i+1, column=3, value=e.get("Created", ""))
                ws.cell(row=i+1, column=4, value=e.get("MessageId", ""))
                ws.cell(row=i+1, column=5, value=e.get("Message", ""))
                # Severity 컬러
                if sev in SEV_FILL:
                    for col in range(1, 6):
                        ws.cell(row=i+1, column=col).fill = SEV_FILL[sev]
                    ws.cell(row=i+1, column=2).font = SEV_FONT[sev]

            # 첫 행 고정
            ws.freeze_panes = "A2"

        wb.save(path)

    # ============================================================
    # 🔧 펌웨어 업데이트
    # ============================================================
    def on_firmware_update(self):
        # 입력 검증
        ip = self.ip_in.text().strip()
        user = self.user_in.text().strip()
        pw = self.pw_in.text()
        missing = []
        if not ip: missing.append("iDRAC IP")
        if not user: missing.append("Username")
        if not pw: missing.append("Password")
        if missing:
            QMessageBox.warning(self, APP_NAME,
                "펌웨어 업데이트에는 iDRAC 접속 정보가 필요합니다:\n• " + "\n• ".join(missing))
            return

        # STEP 1: 위험 경고 + 3개 체크박스
        safety = FirmwareSafetyDialog(self)
        if safety.exec() != QDialog.Accepted:
            return

        # STEP 2: 파일 선택 + 적용 시점
        st = (self.last_payload or {}).get("service_tag", "-")
        target_info = {"ip": ip, "service_tag": st}
        select = FirmwareSelectDialog(self, target_info=target_info)
        if select.exec() != QDialog.Accepted:
            return
        file_path = select.selected_file
        apply_time = select.selected_apply

        # STEP 3: 최종 확인 (UPDATE 타이핑)
        size_mb = os.path.getsize(file_path) / (1024*1024)
        summary = {
            "ip": ip, "service_tag": st,
            "file": file_path, "size_mb": size_mb,
            "apply": "즉시 적용 (자동 재부팅 발생)" if apply_time == "Immediate"
                     else "다음 재부팅 시 적용 (안전)",
        }
        confirm = FirmwareConfirmDialog(self, summary=summary)
        if confirm.exec() != QDialog.Accepted:
            return

        # STEP 4: 진행률 다이얼로그 + Worker 시작
        self.fw_progress_dlg = FirmwareProgressDialog(self)

        self.fw_worker = FirmwareUpdateWorker(ip, user, pw, file_path, apply_time)
        self.fw_worker.stage.connect(self.fw_progress_dlg.set_stage)
        self.fw_worker.upload_progress.connect(self.fw_progress_dlg.set_upload_progress)
        self.fw_worker.install_progress.connect(self.fw_progress_dlg.set_install_progress)
        self.fw_worker.finished_ok.connect(self._on_fw_done)
        self.fw_worker.failed.connect(self._on_fw_fail)
        self.fw_progress_dlg.cancelled.connect(self.fw_worker.cancel)

        self.fw_btn.setEnabled(False)
        self.fw_worker.start()
        self.fw_progress_dlg.exec()

    def _on_fw_done(self, result: dict):
        self.fw_progress_dlg.close()
        self.fw_btn.setEnabled(True)
        reboot_msg = ""
        if result.get("reboot_pending"):
            reboot_msg = ("\n\n📌 <b>다음 재부팅 시</b>에 자동으로 적용됩니다.\n"
                         "관리 대상 서버가 재부팅될 때까지 새 펌웨어는 활성화되지 않습니다.")
        QMessageBox.information(
            self, "업데이트 완료",
            f"펌웨어 업데이트가 성공적으로 완료되었습니다.\n\n"
            f"최종 상태: {result.get('final_state', 'Completed')}\n"
            f"Job URI: {result.get('job_uri', '-')}\n"
            f"메시지: {result.get('message', '')}{reboot_msg}"
        )

    # ============================================================
    # ❓ 문의
    # ============================================================
    def on_contact(self):
        dlg = ContactDialog(self)
        dlg.exec()

    def _on_fw_fail(self, kind: str, msg: str):
        if hasattr(self, 'fw_progress_dlg'):
            self.fw_progress_dlg.close()
        self.fw_btn.setEnabled(True)
        title = {
            "AUTH": "인증 실패", "TIMEOUT": "시간 초과",
            "NETWORK": "네트워크 오류", "UPLOAD": "업로드 실패",
            "JOB": "Job 실패", "BUSY": "iDRAC 사용 중",
            "CANCEL": "취소됨",
        }.get(kind, "오류")
        QMessageBox.critical(self, title, msg)

    def _save_logs_csv(self, path: str, payload: dict, service_tag: str):
        """CSV로 저장 (로그 종류별 별도 파일은 안 함, 하나에 모두)"""
        import csv
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["LogType", "#", "Severity", "Created", "MessageId", "Message"])
            for lt, entries in payload.items():
                lt_name = "LCLog" if lt == "lclog" else "SEL"
                for i, e in enumerate(entries, 1):
                    w.writerow([
                        lt_name, i,
                        e.get("Severity", ""),
                        e.get("Created", ""),
                        e.get("MessageId", ""),
                        e.get("Message", ""),
                    ])


# =========================================================
#  Entry
# =========================================================
def main():
    QApplication.setHighDpiScaleFactorRoundingPolicy(
        Qt.HighDpiScaleFactorRoundingPolicy.PassThrough
    )
    app = QApplication(sys.argv)
    app.setApplicationName(APP_NAME)
    app.setOrganizationName("longchiri")

    # 앱 아이콘 (Dock/Taskbar 까지 적용)
    icon_path = resource_path("iDRAC_Viewer.icns") or resource_path("iDRAC_Viewer.png") or resource_path("iDRAC_Viewer.ico")
    if icon_path:
        app.setWindowIcon(QIcon(icon_path))

    # 전역 스타일시트
    app.setStyleSheet(APP_STYLE)

    # 한글 폰트 fallback
    if sys.platform == "darwin":
        app.setFont(QFont("Apple SD Gothic Neo", 13))
    elif sys.platform.startswith("win"):
        app.setFont(QFont("Malgun Gothic", 10))

    w = MainWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
